"""
Gradebook and grades-related views.
"""
import csv
import logging
from decimal import Decimal, ROUND_HALF_UP
from io import StringIO, BytesIO
from django.db.models import Avg, Sum
from django.http import StreamingHttpResponse, HttpResponse
from django.utils import timezone
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView

logger = logging.getLogger(__name__)

from core.models import (
    Activity,
    AdviserOverrideLog,
    CourseSection,
    Enrollment,
    GradeEntry,
    GradeSubmission,
    GradeSubmissionStatus,
    GradeWeightConfig,
    GradingPeriod,
    Quiz,
    QuizAttempt,
    Section,
    SectionReportCard,
    Submission,
    User,
)
from core.serializers import GradingPeriodSerializer, GradeWeightConfigSerializer
from core.views.common import (
    _letter_grade,
    _compute_enrollment_grade,
)
from core.grade_computation import get_or_create_weight_config


class CourseSectionGradesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        enrollments = list(
            Enrollment.objects.filter(course_section=course_section, is_active=True)
            .select_related("student")
            .order_by("student__last_name", "student__first_name")
        )
        rows = []
        for e in enrollments:
            grade = float(e.final_grade) if e.final_grade is not None else None
            rows.append(
                {
                    "enrollment_id": str(e.id),
                    "student_id": str(e.student_id),
                    "student_name": e.student.full_name,
                    "student_email": e.student.email,
                    "final_grade": grade,
                    "final_grade_letter": _letter_grade(Decimal(str(grade))) if grade is not None else None,
                    "grade_overridden": e.manual_final_grade is not None,
                    "manual_final_grade": float(e.manual_final_grade) if e.manual_final_grade is not None else None,
                }
            )
        return Response(rows)


class CourseSectionGradebookView(APIView):
    """Returns comprehensive gradebook data including activities and quizzes for each student."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Get all published activities and quizzes
        activities = list(
            Activity.objects.filter(course_section=course_section, is_published=True)
            .order_by("created_at", "title")
            .values("id", "title", "points", "deadline", "created_at")
        )
        quizzes = list(
            Quiz.objects.filter(course_section=course_section, is_published=True)
            .order_by("created_at", "title")
            .values("id", "title", "score_selection_policy", "close_at", "created_at")
        )

        # Get quiz max scores
        quiz_points_raw = (
            QuizQuestion.objects.filter(quiz_id__in=[q["id"] for q in quizzes])
            .values("quiz_id")
            .annotate(total=Sum("points"))
        )
        quiz_points = {str(row["quiz_id"]): Decimal(str(row["total"] or 0)) for row in quiz_points_raw}

        # Get all enrollments (both active and inactive)
        all_enrollments = list(
            Enrollment.objects.filter(course_section=course_section)
            .select_related("student")
            .order_by("-is_active", "student__last_name", "student__first_name", "student__email")
        )
        student_ids = [e.student_id for e in all_enrollments]

        # Get all submissions for these activities
        activity_ids = [a["id"] for a in activities]
        submissions = Submission.objects.filter(
            activity_id__in=activity_ids,
            student_id__in=student_ids
        ).values("id", "activity_id", "student_id", "score", "status", "submitted_at", "graded_at")

        # Build submission maps: {(student_id, activity_id): submission_data}
        submission_map: dict[tuple, dict] = {}
        for s in submissions:
            key = (str(s["student_id"]), str(s["activity_id"]))
            submission_map[key] = s

        # Get all quiz attempts
        quiz_ids = [q["id"] for q in quizzes]
        quiz_attempts = QuizAttempt.objects.filter(
            quiz_id__in=quiz_ids,
            student_id__in=student_ids,
            is_submitted=True
        ).values("id", "quiz_id", "student_id", "score", "max_score", "attempt_number", "submitted_at", "pending_manual_grading")

        # Build quiz attempt maps: {(student_id, quiz_id): [attempts]}
        attempts_by_quiz: dict[tuple, list] = {}
        for a in quiz_attempts:
            key = (str(a["student_id"]), str(a["quiz_id"]))
            if key not in attempts_by_quiz:
                attempts_by_quiz[key] = []
            attempts_by_quiz[key].append(a)

        # Build student data
        active_students = []
        inactive_students = []

        for enrollment in all_enrollments:
            student_id = str(enrollment.student_id)
            enrolled_at = enrollment.enrolled_at

            # Build activity grades for this student
            activity_grades = []
            for activity in activities:
                activity_id = str(activity["id"])
                activity_deadline = activity["deadline"]

                # Check if activity deadline is before enrollment (pre-enrollment exclusion)
                if activity_deadline and enrolled_at:
                    from datetime import datetime
                    if isinstance(activity_deadline, str):
                        activity_deadline = datetime.fromisoformat(activity_deadline.replace('Z', '+00:00'))
                    if isinstance(enrolled_at, str):
                        enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
                    else:
                        enrolled_at_dt = enrolled_at
                    if activity_deadline < enrolled_at_dt:
                        activity_grades.append({
                            "activity_id": activity_id,
                            "title": activity["title"],
                            "points": float(activity["points"]),
                            "deadline": activity["deadline"],
                            "score": None,
                            "status": None,
                            "is_late": False,
                            "is_excused": False,
                            "graded_at": None,
                            "is_na": True,  # Pre-enrollment exclusion
                        })
                        continue

                sub = submission_map.get((student_id, activity_id))
                if sub:
                    is_late = False
                    if sub["status"] == Submission.SubmissionStatus.LATE:
                        is_late = True
                    score = float(sub["score"]) if sub["score"] is not None else None
                    graded_at = sub["graded_at"].isoformat() if sub["graded_at"] else None

                    # Determine status
                    status = sub["status"]
                    if score is None and sub["status"] == Submission.SubmissionStatus.SUBMITTED:
                        status = "submitted"  # Needs grading

                    activity_grades.append({
                        "activity_id": activity_id,
                        "title": activity["title"],
                        "points": float(activity["points"]),
                        "deadline": activity["deadline"],
                        "score": score,
                        "status": status if status else "not_submitted",
                        "is_late": is_late,
                        "is_excused": False,
                        "graded_at": graded_at,
                        "is_na": False,
                    })
                else:
                    activity_grades.append({
                        "activity_id": activity_id,
                        "title": activity["title"],
                        "points": float(activity["points"]),
                        "deadline": activity["deadline"],
                        "score": None,
                        "status": "not_submitted",
                        "is_late": False,
                        "is_excused": False,
                        "graded_at": None,
                        "is_na": False,
                    })

            # Build quiz grades for this student
            quiz_grades = []
            for quiz in quizzes:
                quiz_id = str(quiz["id"])
                max_score = float(quiz_points.get(quiz_id, 0))
                close_at = quiz["close_at"]

                # Check if quiz close_at is before enrollment (pre-enrollment exclusion)
                if close_at and enrolled_at:
                    from datetime import datetime
                    if isinstance(close_at, str):
                        close_at_dt = datetime.fromisoformat(close_at.replace('Z', '+00:00'))
                    else:
                        close_at_dt = close_at
                    if isinstance(enrolled_at, str):
                        enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
                    else:
                        enrolled_at_dt = enrolled_at
                    if close_at_dt < enrolled_at_dt:
                        quiz_grades.append({
                            "quiz_id": quiz_id,
                            "title": quiz["title"],
                            "max_score": max_score,
                            "close_at": quiz["close_at"],
                            "score": None,
                            "attempts": 0,
                            "max_attempts": 0,
                            "is_late": False,
                            "is_na": True,
                        })
                        continue

                attempts = attempts_by_quiz.get((student_id, quiz_id), [])
                if attempts:
                    # Calculate score based on policy
                    scores = [float(a["score"]) for a in attempts if a["score"] is not None]
                    if scores:
                        policy = quiz.get("score_selection_policy", "highest")
                        if policy == "highest":
                            score = max(scores)
                        else:  # latest
                            # Sort by attempt_number descending and get the last submitted
                            sorted_attempts = sorted(attempts, key=lambda x: x["attempt_number"], reverse=True)
                            score = float(sorted_attempts[0]["score"]) if sorted_attempts[0]["score"] is not None else None
                    else:
                        score = None

                    # Check if any attempt is pending grading
                    pending_grading = any(a.get("pending_manual_grading", False) for a in attempts)

                    quiz_grades.append({
                        "quiz_id": quiz_id,
                        "title": quiz["title"],
                        "max_score": max_score,
                        "close_at": quiz["close_at"],
                        "score": score,
                        "attempts": len(attempts),
                        "max_attempts": 0,  # TODO: add max_attempts to Quiz model if needed
                        "is_late": False,  # TODO: check if submitted after close_at
                        "is_na": False,
                        "pending_grading": pending_grading,
                    })
                else:
                    quiz_grades.append({
                        "quiz_id": quiz_id,
                        "title": quiz["title"],
                        "max_score": max_score,
                        "close_at": quiz["close_at"],
                        "score": None,
                        "attempts": 0,
                        "max_attempts": 0,
                        "is_late": False,
                        "is_na": False,
                    })

            student_data = {
                "enrollment_id": str(enrollment.id),
                "student_id": student_id,
                "student_name": enrollment.student.full_name,
                "student_email": enrollment.student.email,
                "student_avatar": enrollment.student.avatar_url if hasattr(enrollment.student, 'avatar_url') else None,
                "enrolled_at": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None,
                "is_active": enrollment.is_active,
                "grades": {
                    "activities": activity_grades,
                    "quizzes": quiz_grades,
                },
                "final_grade": float(enrollment.final_grade) if enrollment.final_grade is not None else None,
                "final_grade_letter": _letter_grade(enrollment.final_grade),
                "grade_overridden": enrollment.manual_final_grade is not None,
                "manual_final_grade": float(enrollment.manual_final_grade) if enrollment.manual_final_grade is not None else None,
            }

            if enrollment.is_active:
                active_students.append(student_data)
            else:
                inactive_students.append(student_data)

        # Build items list (column headers)
        items = {
            "activities": [
                {
                    "id": str(a["id"]),
                    "title": a["title"],
                    "type": "activity",
                    "max_points": float(a["points"]),
                    "deadline": a["deadline"].isoformat() if a["deadline"] else None,
                    "created_at": a["created_at"].isoformat() if a["created_at"] else None,
                }
                for a in activities
            ],
            "quizzes": [
                {
                    "id": str(q["id"]),
                    "title": q["title"],
                    "type": "quiz",
                    "max_points": float(quiz_points.get(str(q["id"]), 0)),
                    "close_at": q["close_at"].isoformat() if q["close_at"] else None,
                    "created_at": q["created_at"].isoformat() if q["created_at"] else None,
                }
                for q in quizzes
            ],
        }

        # Build summary statistics for each item
        activity_summary = []
        for activity in activities:
            activity_id = str(activity["id"])
            scores = []
            missing_count = 0
            needs_grading_count = 0

            for student in active_students:
                grade = next((g for g in student["grades"]["activities"] if g["activity_id"] == activity_id), None)
                if grade:
                    if grade.get("is_na"):
                        continue  # Exclude from stats
                    if grade["score"] is not None:
                        scores.append(grade["score"])
                    elif grade["status"] == "not_submitted":
                        missing_count += 1
                    elif grade["status"] == "submitted":
                        needs_grading_count += 1
                    elif grade["status"] == "late":
                        needs_grading_count += 1  # Late submissions also need grading if no score

            activity_summary.append({
                "activity_id": activity_id,
                "avg_score": float(sum(scores) / len(scores)) if scores else None,
                "high_score": float(max(scores)) if scores else None,
                "low_score": float(min(scores)) if scores else None,
                "missing_count": missing_count,
                "needs_grading_count": needs_grading_count,
            })

        quiz_summary = []
        for quiz in quizzes:
            quiz_id = str(quiz["id"])
            scores = []
            missing_count = 0
            needs_grading_count = 0

            for student in active_students:
                grade = next((g for g in student["grades"]["quizzes"] if g["quiz_id"] == quiz_id), None)
                if grade:
                    if grade.get("is_na"):
                        continue
                    if grade["score"] is not None:
                        scores.append(grade["score"])
                    elif grade["attempts"] == 0:
                        missing_count += 1
                    if grade.get("pending_grading"):
                        needs_grading_count += 1

            quiz_summary.append({
                "quiz_id": quiz_id,
                "avg_score": float(sum(scores) / len(scores)) if scores else None,
                "high_score": float(max(scores)) if scores else None,
                "low_score": float(min(scores)) if scores else None,
                "missing_count": missing_count,
                "needs_grading_count": needs_grading_count,
            })

        return Response({
            "students": active_students,
            "inactive_students": inactive_students,
            "items": items,
            "summary": {
                "activities": activity_summary,
                "quizzes": quiz_summary,
            },
        })


class CourseSectionGradesExportCSVView(APIView):
    """Export grades as CSV or XLSX with optional filtering."""
    permission_classes = [permissions.IsAuthenticated]

    def _format_num(self, value: Decimal | float | int | None) -> str:
        if value is None:
            return ""
        dec = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{dec:.2f}"

    def _stream_text(self, text: str, chunk_size: int = 8192):
        for idx in range(0, len(text), chunk_size):
            yield text[idx : idx + chunk_size]

    def _resolve_course_section(self, *, pk=None, course_id=None, section_id=None):
        from core.models import CourseSection
        target_id = section_id or pk
        if not target_id:
            return None
        course_section = CourseSection.objects.select_related("course", "section").filter(id=target_id).first()
        if not course_section:
            return None
        if course_id and str(course_section.course_id) != str(course_id):
            return None
        return course_section

    def _generate_csv(self, course_section, enrollments, activities, quizzes, submission_map, quiz_attempt_map, include_inactive=False, include_student_id=False, include_enrolled_at=False):
        """Generate CSV export."""
        headers = ["Student Name", "Student Email"]
        if include_student_id:
            headers.append("Student ID")
        if include_enrolled_at:
            headers.append("Enrolled At")
        headers.extend(["Section", "Total Grade", "Grade Letter"])
        headers.extend([f"Activity: {a.title}" for a in activities])
        headers.extend([f"Quiz: {q.title}" for q in quizzes])

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)

        for enrollment in enrollments:
            if not enrollment.is_active and not include_inactive:
                continue
            computed_total = enrollment.final_grade
            row = [
                enrollment.student.full_name,
                enrollment.student.email,
            ]
            if include_student_id:
                row.append(str(enrollment.student_id))
            if include_enrolled_at:
                row.append(enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else "")
            row.extend([
                course_section.section.name,
                self._format_num(computed_total),
                _letter_grade(computed_total) if computed_total else "",
            ])
            for activity in activities:
                sub = submission_map.get((str(enrollment.student_id), str(activity.id)))
                if sub and sub.score is not None:
                    row.append(f"{self._format_num(sub.score)}/{self._format_num(activity.points)}")
                else:
                    row.append("")
            for quiz in quizzes:
                attempt = quiz_attempt_map.get((str(enrollment.student_id), str(quiz.id)))
                if attempt and attempt.score is not None and attempt.max_score is not None:
                    row.append(f"{self._format_num(attempt.score)}/{self._format_num(attempt.max_score)}")
                elif attempt and attempt.score is not None:
                    row.append(self._format_num(attempt.score))
                else:
                    row.append("")
            writer.writerow(row)

        return "\ufeff" + output.getvalue()

    def _generate_xlsx(self, course_section, enrollments, activities, quizzes, submission_map, quiz_attempt_map, include_inactive=False, include_student_id=False, include_enrolled_at=False):
        """Generate XLSX export with conditional formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"

        # Define styles
        header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        missing_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red
        needs_grading_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        passing_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Build headers
        headers = ["Student Name", "Student Email"]
        if include_student_id:
            headers.append("Student ID")
        if include_enrolled_at:
            headers.append("Enrolled At")
        headers.extend(["Section", "Total Grade", "Grade Letter"])
        headers.extend([f"Activity: {a.title}" for a in activities])
        headers.extend([f"Quiz: {q.title}" for q in quizzes])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 25  # Student Name
        ws.column_dimensions['B'].width = 30  # Student Email
        if include_student_id:
            ws.column_dimensions['C'].width = 36  # Student ID
        if include_enrolled_at:
            extra_col = 1 if include_student_id else 0
            ws.column_dimensions[get_column_letter(3 + extra_col)].width = 20  # Enrolled At

        # Data rows
        row_num = 2
        for enrollment in enrollments:
            if not enrollment.is_active and not include_inactive:
                continue
            computed_total = enrollment.final_grade

            col = 1
            ws.cell(row=row_num, column=col, value=enrollment.student.full_name).border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=enrollment.student.email).border = thin_border
            col += 1
            if include_student_id:
                ws.cell(row=row_num, column=col, value=str(enrollment.student_id)).border = thin_border
                col += 1
            if include_enrolled_at:
                ws.cell(row=row_num, column=col, value=enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else "").border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value=course_section.section.name).border = thin_border
            col += 1
            grade_cell = ws.cell(row=row_num, column=col, value=float(computed_total) if computed_total else "")
            grade_cell.border = thin_border
            if computed_total and computed_total >= Decimal("70"):
                grade_cell.fill = passing_fill
            col += 1
            ws.cell(row=row_num, column=col, value=_letter_grade(computed_total) if computed_total else "").border = thin_border
            col += 1

            # Activity scores
            for activity in activities:
                sub = submission_map.get((str(enrollment.student_id), str(activity.id)))
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                if sub and sub.score is not None:
                    cell.value = f"{float(sub.score):.1f}/{float(activity.points):.0f}"
                    if sub.score / activity.points >= Decimal("0.7"):
                        cell.fill = passing_fill
                else:
                    cell.value = ""
                    cell.fill = missing_fill
                col += 1

            # Quiz scores
            for quiz in quizzes:
                attempt = quiz_attempt_map.get((str(enrollment.student_id), str(quiz.id)))
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                if attempt and attempt.score is not None and attempt.max_score is not None:
                    cell.value = f"{float(attempt.score):.1f}/{float(attempt.max_score):.0f}"
                    if attempt.score / attempt.max_score >= Decimal("0.7"):
                        cell.fill = passing_fill
                elif attempt and attempt.score is not None:
                    cell.value = f"{float(attempt.score):.1f}"
                else:
                    cell.value = ""
                    cell.fill = missing_fill
                col += 1

            row_num += 1

        # Freeze first row (header)
        ws.freeze_panes = 'A2'

        # Write to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def get(self, request, pk=None, course_id=None, section_id=None):
        course_section = self._resolve_course_section(pk=pk, course_id=course_id, section_id=section_id)
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Query parameters
        export_format = request.query_params.get("format", "csv").lower()
        scope = request.query_params.get("scope", "all").lower()  # all, activities, quizzes, final_only
        include_inactive = request.query_params.get("include_inactive", "false").lower() == "true"
        include_student_id = request.query_params.get("include_student_id", "false").lower() == "true"
        include_enrolled_at = request.query_params.get("include_enrolled_at", "false").lower() == "true"

        activities = list(
            Activity.objects.filter(course_section=course_section, is_published=True).order_by("created_at", "title")
        )
        quizzes = list(
            Quiz.objects.filter(course_section=course_section, is_published=True).order_by("created_at", "title")
        )

        # Filter by scope
        if scope == "activities":
            quizzes = []
        elif scope == "quizzes":
            activities = []
        elif scope == "final_only":
            activities = []
            quizzes = []

        enrollments = list(
            Enrollment.objects.filter(course_section=course_section)
            .select_related("student", "course_section")
            .order_by("-is_active", "student__last_name", "student__first_name", "student__email")
        )
        student_ids = [e.student_id for e in enrollments]

        submission_map: dict[tuple[str, str], Submission] = {}
        if activities and student_ids:
            submissions = Submission.objects.filter(
                activity_id__in=[a.id for a in activities],
                student_id__in=student_ids,
            )
            submission_map = {(str(s.student_id), str(s.activity_id)): s for s in submissions}

        quiz_attempt_map: dict[tuple[str, str], QuizAttempt] = {}
        if quizzes and student_ids:
            quiz_attempts = (
                QuizAttempt.objects.filter(
                    quiz_id__in=[q.id for q in quizzes],
                    student_id__in=student_ids,
                    is_submitted=True,
                    score__isnull=False,
                )
                .order_by("student_id", "quiz_id", "-attempt_number", "-submitted_at")
            )
            for attempt in quiz_attempts:
                key = (str(attempt.student_id), str(attempt.quiz_id))
                if key not in quiz_attempt_map:
                    quiz_attempt_map[key] = attempt

        base_name = f"{course_section.course.code}_{course_section.section.name}_grades".replace(" ", "_")

        if export_format == "xlsx":
            xlsx_data = self._generate_xlsx(
                course_section, enrollments, activities, quizzes,
                submission_map, quiz_attempt_map,
                include_inactive, include_student_id, include_enrolled_at
            )
            if xlsx_data is None:
                return Response({"detail": "XLSX generation not available."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            response = HttpResponse(
                xlsx_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{base_name}.xlsx"'
            return response

        # Default to CSV
        csv_text = self._generate_csv(
            course_section, enrollments, activities, quizzes,
            submission_map, quiz_attempt_map,
            include_inactive, include_student_id, include_enrolled_at
        )

        import logging
        logging.getLogger(__name__).info(
            "grade_csv_export user_id=%s section_id=%s students=%s activities=%s quizzes=%s",
            request.user.id,
            course_section.id,
            len([e for e in enrollments if e.is_active]),
            len(activities),
            len(quizzes),
        )

        response = StreamingHttpResponse(self._stream_text(csv_text), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
        return response


class EnrollmentGradeOverrideView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        enrollment = Enrollment.objects.select_related("course_section").filter(id=pk).first()
        if not enrollment:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and enrollment.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        value = request.data.get("manual_final_grade", None)
        if value in ["", None]:
            enrollment.manual_final_grade = None
        else:
            try:
                enrollment.manual_final_grade = Decimal(str(value))
            except Exception:
                return Response({"detail": "manual_final_grade must be numeric."}, status=status.HTTP_400_BAD_REQUEST)

        enrollment.save(update_fields=["manual_final_grade"])
        enrollment.final_grade = _compute_enrollment_grade(enrollment)
        return Response(
            {
                "enrollment_id": str(enrollment.id),
                "manual_final_grade": float(enrollment.manual_final_grade) if enrollment.manual_final_grade is not None else None,
                "final_grade": float(enrollment.final_grade) if enrollment.final_grade is not None else None,
                "final_grade_letter": _letter_grade(enrollment.final_grade),
                "grade_overridden": enrollment.manual_final_grade is not None,
            }
        )


# Import QuizQuestion for use in CourseSectionGradebookView
from core.models import QuizQuestion


class GradingPeriodListView(APIView):
    """List grading periods, optionally filtered by school_year."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        school_year = request.query_params.get('school_year')
        queryset = GradingPeriod.objects.all()
        if school_year:
            queryset = queryset.filter(school_year=school_year)
        queryset = queryset.order_by('school_year', 'period_number')
        serializer = GradingPeriodSerializer(queryset, many=True)
        return Response(serializer.data)


class StudentGradesView(APIView):
    """Return a student's own grades for a course section (published entries only)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        from core.models import CourseSection, TeacherAdvisory

        # Get the course section
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Only students can access this view for their own grades
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can view their own grades."}, status=status.HTTP_403_FORBIDDEN)

        # Get the student's enrollment
        enrollment = Enrollment.objects.filter(
            course_section=course_section,
            student=request.user,
            is_active=True
        ).first()
        if not enrollment:
            return Response({"detail": "You are not enrolled in this course."}, status=status.HTTP_404_NOT_FOUND)

        # Determine semester_group filter based on grade level and course section semester
        section = course_section.section
        grade_level_str = section.grade_level if section else None

        if grade_level_str in ['Grade 11', 'Grade 12']:
            # For Senior High, filter by semester_group based on course_section.semester
            semester_value = course_section.semester
            if semester_value in ['1st', '1', 'First']:
                semester_group = 1
            elif semester_value in ['2nd', '2', 'Second']:
                semester_group = 2
            else:
                semester_group = None
        else:
            semester_group = None  # Grades 7-10 have no semester grouping

        # Get grading periods for this school year
        school_year = course_section.school_year
        if semester_group is not None:
            grading_periods = GradingPeriod.objects.filter(
                school_year=school_year,
                semester_group=semester_group
            ).order_by('period_number')
        else:
            grading_periods = GradingPeriod.objects.filter(
                school_year=school_year,
                semester_group__isnull=True
            ).order_by('period_number')

        # Get grade entries for this enrollment
        grade_entries = GradeEntry.objects.filter(
            enrollment=enrollment,
            is_published=True  # Only published entries for students
        ).select_related('grading_period')

        # Build period grades
        entry_by_period = {str(e.grading_period_id): e for e in grade_entries}
        period_grades = []
        for period in grading_periods:
            entry = entry_by_period.get(str(period.id))
            period_grades.append({
                'period': GradingPeriodSerializer(period).data,
                'score': float(entry.score) if entry and entry.score is not None else None,
                'is_published': entry.is_published if entry else False,
            })

        # Get final grade from enrollment
        final_grade = enrollment.final_grade
        final_grade_letter = _letter_grade(Decimal(str(final_grade))) if final_grade else None

        # Deprecation warning
        logger.warning(
            'DEPRECATED endpoint called: course-sections grades student view. '
            'Caller: %s', request.META.get('HTTP_USER_AGENT', 'unknown')
        )

        # Build response
        response = Response({
            'course_section_id': str(course_section.id),
            'course_code': course_section.course.code,
            'course_title': course_section.course.title,
            'grade_level': grade_level_str,
            'periods': period_grades,
            'final_grade': float(final_grade) if final_grade and enrollment.is_final_published else None,
            'final_grade_letter': final_grade_letter if enrollment.is_final_published else None,
            'is_final_published': enrollment.is_final_published,
        })
        response['X-Deprecated'] = (
            'This endpoint is deprecated. '
            'Use GET /api/students/me/report-card/ instead. '
            'Will be removed after React Native app migration.'
        )
        return response


class AdvisoryGradesView(APIView):
    """Return all students' grades across all subjects for an advisory section."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, section_id):
        from core.models import CourseSection, TeacherAdvisory, Section

        # Get the section directly
        section = Section.objects.filter(id=section_id).first()
        if not section:
            return Response({"detail": "Section not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get the active advisory assignment to determine school year
        advisory = TeacherAdvisory.objects.filter(
            section=section,
            is_active=True
        ).select_related('section').first()
        if not advisory:
            return Response({"detail": "No active advisory for this section."}, status=status.HTTP_404_NOT_FOUND)

        school_year = advisory.school_year

        # Check if user is an advisory teacher for this section
        is_adviser = TeacherAdvisory.objects.filter(
            teacher=request.user,
            section=section,
            school_year=school_year,
            is_active=True
        ).exists()

        # Admins can also access
        if request.user.role != User.Role.ADMIN and not is_adviser:
            return Response({"detail": "You are not the advisory teacher for this section."}, status=status.HTTP_403_FORBIDDEN)

        # Get all students enrolled in this section (via enrollments in all course_sections for this section)
        enrollments_in_section = Enrollment.objects.filter(
            course_section__section=section,
            course_section__school_year=school_year,
            is_active=True
        ).select_related('student', 'course_section__course')

        # Get all unique students
        students_dict = {}  # student_id -> student data
        for enrollment in enrollments_in_section:
            student_id = str(enrollment.student.id)
            if student_id not in students_dict:
                students_dict[student_id] = {
                    'student_id': student_id,
                    'student_name': enrollment.student.get_full_name(),
                    'student_email': enrollment.student.email,
                    'subjects': [],
                    'final_average': None,
                }

        # Determine semester_group filter based on grade level
        # Grades 7-10: All quarters (semester_group is null)
        # Grades 11-12: Quarters grouped by semester
        grade_level = section.grade_level if section else None
        if grade_level in ['Grade 11', 'Grade 12']:
            # For Senior High, show all quarters for this school year
            # The advisory teacher sees grades from all subjects across both semesters
            grading_periods = list(GradingPeriod.objects.filter(
                school_year=school_year,
                semester_group__isnull=False
            ).order_by('semester_group', 'period_number'))
        else:
            # For Grades 7-10, show all quarters
            grading_periods = list(GradingPeriod.objects.filter(
                school_year=school_year,
                semester_group__isnull=True
            ).order_by('period_number'))

        # Fallback: If no grading periods found for advisory's school_year,
        # try to use the most recent school_year from existing GradingPeriods.
        # This handles cases where a new school_year is set but GradingPeriods haven't been created yet.
        if not grading_periods:
            latest_period = GradingPeriod.objects.order_by('-school_year').first()
            if latest_period:
                fallback_year = latest_period.school_year
                logger.warning(
                    f"No grading periods found for school_year '{school_year}' "
                    f"in AdvisoryGradesView for section {section_id}. "
                    f"Falling back to most recent school_year '{fallback_year}'."
                )
                school_year = fallback_year
                if grade_level in ['Grade 11', 'Grade 12']:
                    grading_periods = list(GradingPeriod.objects.filter(
                        school_year=school_year,
                        semester_group__isnull=False
                    ).order_by('semester_group', 'period_number'))
                else:
                    grading_periods = list(GradingPeriod.objects.filter(
                        school_year=school_year,
                        semester_group__isnull=True
                    ).order_by('period_number'))

        # Get all grade entries for these enrollments
        # Advisers see all grades (not just published) for override capability
        enrollment_ids = [str(e.id) for e in enrollments_in_section]
        grade_entries = GradeEntry.objects.filter(
            enrollment_id__in=enrollment_ids,
        ).select_related('enrollment', 'grading_period')

        # Group grade entries by enrollment
        entries_by_enrollment = {}
        for entry in grade_entries:
            enrollment_id = str(entry.enrollment_id)
            if enrollment_id not in entries_by_enrollment:
                entries_by_enrollment[enrollment_id] = []
            entries_by_enrollment[enrollment_id].append(entry)

        # Build subject grades for each student
        subject_dict = {}  # course_section_id -> course info
        for enrollment in enrollments_in_section:
            student_id = str(enrollment.student.id)
            course_section_id = str(enrollment.course_section_id)

            # Track unique subjects
            if course_section_id not in subject_dict:
                course = enrollment.course_section.course
                teacher_name = enrollment.course_section.teacher.get_full_name() if enrollment.course_section.teacher else 'Unassigned'
                subject_dict[course_section_id] = {
                    'subject_id': course_section_id,
                    'subject_code': course.code,
                    'subject_title': course.title,
                    'teacher_name': teacher_name,
                }

            # Get grades for this enrollment
            entries = entries_by_enrollment.get(str(enrollment.id), [])
            entry_by_period = {str(e.grading_period_id): e for e in entries}

            period_grades = []
            for period in grading_periods:
                entry = entry_by_period.get(str(period.id))
                period_grades.append({
                    'grading_period_id': str(period.id),
                    'period_label': period.label,
                    'score': float(entry.score) if entry and entry.score is not None else None,
                    'is_published': entry.is_published if entry else False,
                    'grade_entry_id': str(entry.id) if entry else None,
                })

            subject_grade = {
                'subject_id': course_section_id,
                'subject_code': enrollment.course_section.course.code,
                'subject_title': enrollment.course_section.course.title,
                'teacher_name': subject_dict[course_section_id]['teacher_name'],
                'periods': period_grades,
                'final_grade': float(enrollment.final_grade) if enrollment.final_grade else None,
            }

            if student_id in students_dict:
                students_dict[student_id]['subjects'].append(subject_grade)

        # Calculate final averages (average of final_grades across subjects)
        for student_id, student_data in students_dict.items():
            final_grades = [s['final_grade'] for s in student_data['subjects'] if s['final_grade'] is not None]
            if final_grades:
                student_data['final_average'] = round(sum(final_grades) / len(final_grades), 2)

        # Get submission status per subject (course_section) for each grading period
        course_section_ids = list(subject_dict.keys())
        submissions = GradeSubmission.objects.filter(
            course_section_id__in=course_section_ids,
            grading_period__in=grading_periods,
        ).select_related('course_section', 'grading_period')

        submission_status = []
        for cs_id, info in subject_dict.items():
            cs_submissions = [s for s in submissions if str(s.course_section_id) == cs_id]
            submission_status.append({
                'subject_id': cs_id,
                'subject_code': info['subject_code'],
                'subject_title': info['subject_title'],
                'teacher_name': info.get('teacher_name', ''),
                'periods': [
                    {
                        'grading_period_id': str(gp.id),
                        'period_label': gp.label,
                        'status': next(
                            (s.status for s in cs_submissions if str(s.grading_period_id) == str(gp.id)),
                            GradeSubmissionStatus.DRAFT,
                        ),
                    }
                    for gp in grading_periods
                ],
            })

        # Get report card status per period
        report_cards = SectionReportCard.objects.filter(
            section=section,
            grading_period__in=grading_periods,
        ).select_related('grading_period')

        report_card_by_period = {str(rc.grading_period_id): rc for rc in report_cards}
        report_card_status = []
        for gp in grading_periods:
            rc = report_card_by_period.get(str(gp.id))
            report_card_status.append({
                'grading_period_id': str(gp.id),
                'period_label': gp.label,
                'is_published': rc.is_published if rc else False,
                'published_at': rc.published_at if rc else None,
            })

        # Build response
        students_list = list(students_dict.values())
        # Sort by student name
        students_list.sort(key=lambda x: x['student_name'])

        return Response({
            'section_id': str(section.id),
            'section_name': section.name,
            'grade_level': section.grade_level,
            'strand': section.strand,
            'school_year': school_year,
            'students': students_list,
            'periods': GradingPeriodSerializer(grading_periods, many=True).data,
            'submission_status': submission_status,
            'report_card_status': report_card_status,
        })


class SubjectGradesView(APIView):
    """Return subject teacher's gradebook with per-period breakdown."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions - teacher must be assigned to this course section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Get all enrollments
        enrollments = list(
            Enrollment.objects.filter(course_section=course_section, is_active=True)
            .select_related('student')
            .order_by('student__last_name', 'student__first_name')
        )

        # Get grade level from section to determine period type
        grade_level = None
        if course_section.section:
            grade_level = course_section.section.grade_level

        # Determine period type based on grade level
        # Grades 7-10 use quarters, Grades 11-12 use semesters
        if grade_level in ['Grade 11', 'Grade 12']:
            # For Senior High, filter by semester_group based on course_section.semester
            # semester='1st' -> Q1, Q2 (semester_group=1)
            # semester='2nd' -> Q3, Q4 (semester_group=2)
            semester_value = course_section.semester
            if semester_value in ['1st', '1', 'First']:
                semester_group = 1
            elif semester_value in ['2nd', '2', 'Second']:
                semester_group = 2
            else:
                semester_group = None  # Show all if semester not set

            if semester_group:
                grading_periods = list(GradingPeriod.objects.filter(
                    school_year=course_section.school_year,
                    semester_group=semester_group
                ).order_by('period_number'))
            else:
                # Fallback: show all quarters for this school year
                grading_periods = list(GradingPeriod.objects.filter(
                    school_year=course_section.school_year
                ).order_by('semester_group', 'period_number'))
        else:
            # For Grades 7-10, show all quarters (semester_group is null)
            grading_periods = list(GradingPeriod.objects.filter(
                school_year=course_section.school_year,
                semester_group__isnull=True
            ).order_by('period_number'))

        # Get all grade entries for these enrollments
        enrollment_ids = [str(e.id) for e in enrollments]
        grade_entries = GradeEntry.objects.filter(
            enrollment_id__in=enrollment_ids
        ).select_related('enrollment', 'grading_period')

        # Group by enrollment
        entries_by_enrollment = {}
        for entry in grade_entries:
            enrollment_id = str(entry.enrollment_id)
            if enrollment_id not in entries_by_enrollment:
                entries_by_enrollment[enrollment_id] = {}
            entries_by_enrollment[enrollment_id][str(entry.grading_period_id)] = entry

        # Build student data
        students = []
        for enrollment in enrollments:
            entry_map = entries_by_enrollment.get(str(enrollment.id), {})

            period_grades = []
            for period in grading_periods:
                entry = entry_map.get(str(period.id))
                period_grades.append({
                    'period_id': str(period.id),
                    'period_label': period.label,
                    'grade_entry_id': str(entry.id) if entry else None,
                    'score': float(entry.score) if entry and entry.score is not None else None,
                    'computed_score': float(entry.computed_score) if entry and entry.computed_score is not None else None,
                    'override_score': float(entry.override_score) if entry and entry.override_score is not None else None,
                    'is_published': entry.is_published if entry else False,
                })

            students.append({
                'enrollment_id': str(enrollment.id),
                'student_id': str(enrollment.student.id),
                'student_name': enrollment.student.get_full_name(),
                'student_email': enrollment.student.email,
                'periods': period_grades,
                'final_grade': float(enrollment.final_grade) if enrollment.final_grade else None,
                'final_grade_letter': _letter_grade(enrollment.final_grade) if enrollment.final_grade else None,
                'grade_overridden': enrollment.manual_final_grade is not None,
                'is_final_published': enrollment.is_final_published,
            })

        # Check if all final grades are published
        all_final_published = all(s.get('is_final_published', False) for s in students) if students else False

        # Get submission status for each grading period
        submissions = GradeSubmission.objects.filter(
            course_section=course_section,
            grading_period__in=grading_periods,
        ).select_related('grading_period')

        submission_by_period = {str(s.grading_period_id): s for s in submissions}
        submission_list = []
        for period in grading_periods:
            sub = submission_by_period.get(str(period.id))
            submission_list.append({
                'grading_period_id': str(period.id),
                'period_label': period.label,
                'status': sub.status if sub else GradeSubmissionStatus.DRAFT,
                'submitted_by': str(sub.submitted_by_id) if sub and sub.submitted_by_id else None,
                'submitted_at': sub.submitted_at if sub else None,
            })

        return Response({
            'course_section_id': str(course_section.id),
            'course_code': course_section.course.code,
            'course_title': course_section.course.title,
            'grade_level': grade_level,
            'semester_group': semester_group,  # For Grades 11-12, which semester this course belongs to
            'periods': GradingPeriodSerializer(grading_periods, many=True).data,
            'submissions': submission_list,
            'students': students,
            'all_final_published': all_final_published,
        })


class GradeEntryUpdateView(APIView):
    """Update a grade entry's override score."""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        entry = GradeEntry.objects.select_related('enrollment__course_section').filter(id=pk).first()
        if not entry:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions - teacher must be assigned to this course section
        course_section = entry.enrollment.course_section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Check GradeSubmission status - only allow edits when status is draft
        submission = GradeSubmission.objects.filter(
            course_section=course_section,
            grading_period=entry.grading_period,
        ).first()
        if submission and submission.status != GradeSubmissionStatus.DRAFT:
            return Response(
                {"detail": f"Cannot edit grades: submission status is {submission.status}. Only draft submissions can be edited."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update override score
        override_score = request.data.get('override_score')
        if override_score is None or override_score == '':
            entry.override_score = None
        else:
            try:
                entry.override_score = Decimal(str(override_score))
            except Exception:
                return Response({"detail": "override_score must be numeric."}, status=status.HTTP_400_BAD_REQUEST)

        entry.save()

        return Response({
            'id': str(entry.id),
            'score': float(entry.score) if entry.score is not None else None,
            'computed_score': float(entry.computed_score) if entry.computed_score is not None else None,
            'override_score': float(entry.override_score) if entry.override_score is not None else None,
        })


class GradeEntryPublishView(APIView):
    """Toggle publish status for a grade entry."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        entry = GradeEntry.objects.select_related('enrollment__course_section').filter(id=pk).first()
        if not entry:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions - teacher must be assigned to this course section
        course_section = entry.enrollment.course_section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Toggle publish status
        is_published = request.data.get('is_published', not entry.is_published)
        entry.is_published = bool(is_published)
        entry.save()

        return Response({
            'id': str(entry.id),
            'is_published': entry.is_published,
        })


class BulkPublishGradesView(APIView):
    """Bulk publish all grades for a grading period in a course section."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        # pk is course_section_id
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions - teacher must be assigned to this course section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        grading_period_id = request.data.get("grading_period_id")
        if not grading_period_id:
            return Response({"detail": "grading_period_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Update all grade entries
        from core.models import GradeEntry
        updated_count = GradeEntry.objects.filter(
            enrollment__course_section=course_section,
            grading_period_id=grading_period_id
        ).update(is_published=True)

        return Response({
            'published_count': updated_count,
        })


def _compute_final_grade_from_entries(enrollment: Enrollment, semester_group: int | None = None) -> Decimal | None:
    """
    Compute final grade from grade entries.

    For Grades 7-10: Average of all 4 quarters
    For Grades 11-12: Average of 2 quarters in the semester

    Returns None if no published grade entries exist.
    """
    grade_level = enrollment.course_section.course.grade_level if enrollment.course_section.course else None

    # Filter grade entries by semester_group for SHS
    if semester_group is not None:
        # Get grading periods for this semester
        period_ids = GradingPeriod.objects.filter(
            semester_group=semester_group,
            school_year=enrollment.course_section.school_year
        ).values_list('id', flat=True)

        entries = GradeEntry.objects.filter(
            enrollment=enrollment,
            is_published=True,
            grading_period_id__in=period_ids
        )
    else:
        # All quarters for JHS
        entries = GradeEntry.objects.filter(
            enrollment=enrollment,
            is_published=True
        )

    if not entries.exists():
        return None

    scores = [e.score for e in entries if e.score is not None]

    if not scores:
        return None

    # Calculate average
    total = sum(scores)
    average = Decimal(str(total)) / Decimal(str(len(scores)))

    # Round to 2 decimal places
    return average.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


class ComputeFinalGradeView(APIView):
    """Compute and optionally publish final grade from grade entries."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        """
        Compute final grade for an enrollment.

        Request body:
        - publish: bool - Whether to publish the final grade
        """
        from core.models import Enrollment

        enrollment = Enrollment.objects.filter(id=pk).first()
        if not enrollment:
            return Response({"detail": "Enrollment not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        course_section = enrollment.course_section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Determine semester_group
        grade_level = course_section.course.grade_level if course_section.course else None
        semester_group = None
        if grade_level in ['Grade 11', 'Grade 12']:
            semester_value = course_section.semester
            if semester_value in ['1st', '1', 'First']:
                semester_group = 1
            elif semester_value in ['2nd', '2', 'Second']:
                semester_group = 2

        # Compute final grade from entries
        final_grade = _compute_final_grade_from_entries(enrollment, semester_group)

        if final_grade is None:
            return Response({
                "detail": "No published grade entries found.",
                "final_grade": None,
            })

        # Check if we should publish
        publish = request.data.get('publish', False)
        if publish:
            enrollment.final_grade = final_grade
            enrollment.is_final_published = True
            enrollment.save(update_fields=['final_grade', 'is_final_published'])
        else:
            # Just return the computed value without saving
            pass

        return Response({
            'enrollment_id': str(enrollment.id),
            'final_grade': float(final_grade),
            'final_grade_letter': _letter_grade(final_grade),
            'is_published': enrollment.is_final_published if publish else False,
        })


class BulkPublishFinalGradesView(APIView):
    """Bulk publish final grades for all students in a course section."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        """
        Compute and publish final grades for all enrollments in a course section.

        Final grades are computed from grade entries:
        - Grades 7-10: Average of all 4 quarters
        - Grades 11-12: Average of 2 quarters in the semester
        """
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Determine semester_group
        grade_level = course_section.course.grade_level if course_section.course else None
        semester_group = None
        if grade_level in ['Grade 11', 'Grade 12']:
            semester_value = course_section.semester
            if semester_value in ['1st', '1', 'First']:
                semester_group = 1
            elif semester_value in ['2nd', '2', 'Second']:
                semester_group = 2

        # Get all active enrollments
        enrollments = Enrollment.objects.filter(
            course_section=course_section,
            is_active=True
        )

        published_count = 0
        for enrollment in enrollments:
            # Compute final grade from entries
            final_grade = _compute_final_grade_from_entries(enrollment, semester_group)

            if final_grade is not None:
                enrollment.final_grade = final_grade
                enrollment.is_final_published = True
                enrollment.save(update_fields=['final_grade', 'is_final_published'])
                published_count += 1

        return Response({
            'published_count': published_count,
        })


class GradeWeightConfigView(APIView):
    """GET/PUT grade weight configuration for a course section."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.select_related('course').filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Permission check: subject teacher or admin
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        config = get_or_create_weight_config(course_section)
        serializer = GradeWeightConfigSerializer(config)
        return Response(serializer.data)

    def put(self, request, pk):
        course_section = CourseSection.objects.select_related('course').filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Permission check: subject teacher or admin
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        written_works = request.data.get('written_works')
        performance_tasks = request.data.get('performance_tasks')
        quarterly_assessment = request.data.get('quarterly_assessment')

        # Validate required fields
        if written_works is None or performance_tasks is None or quarterly_assessment is None:
            return Response(
                {"detail": "written_works, performance_tasks, and quarterly_assessment are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate they are integers
        try:
            written_works = int(written_works)
            performance_tasks = int(performance_tasks)
            quarterly_assessment = int(quarterly_assessment)
        except (ValueError, TypeError):
            return Response(
                {"detail": "All weights must be integers."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate sum equals 100
        total = written_works + performance_tasks + quarterly_assessment
        if total != 100:
            return Response(
                {"detail": f"Weights must sum to 100. Current total: {total}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get or create config, then update
        config = get_or_create_weight_config(course_section)
        config.written_works = written_works
        config.performance_tasks = performance_tasks
        config.quarterly_assessment = quarterly_assessment
        config.is_customized = True
        config.updated_by = request.user
        config.save()

        serializer = GradeWeightConfigSerializer(config)
        return Response(serializer.data)


class GradeSubmissionSubmitView(APIView):
    """Subject teacher submits period grades."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        # pk = course_section_id
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Only subject teacher or admin can submit
        if request.user.role != User.Role.ADMIN and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        grading_period_id = request.data.get("grading_period_id")
        if not grading_period_id:
            return Response({"detail": "grading_period_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        grading_period = GradingPeriod.objects.filter(id=grading_period_id).first()
        if not grading_period:
            return Response({"detail": "Grading period not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check that at least one grade entry exists
        has_entries = GradeEntry.objects.filter(
            enrollment__course_section=course_section,
            grading_period=grading_period,
        ).exists()
        if not has_entries:
            return Response({"detail": "No grade entries exist for this period."}, status=status.HTTP_400_BAD_REQUEST)

        # Get or create the submission
        submission, created = GradeSubmission.objects.get_or_create(
            course_section=course_section,
            grading_period=grading_period,
        )

        # Can only submit from draft status
        if submission.status != GradeSubmissionStatus.DRAFT:
            return Response({"detail": f"Cannot submit: current status is {submission.status}."}, status=status.HTTP_400_BAD_REQUEST)

        submission.status = GradeSubmissionStatus.SUBMITTED
        submission.submitted_by = request.user
        submission.submitted_at = timezone.now()
        submission.save()

        return Response({
            "submission_id": str(submission.id),
            "status": submission.status,
            "submitted_at": submission.submitted_at,
        })


class GradeSubmissionTakeBackView(APIView):
    """Subject teacher takes back submitted period grades."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role != User.Role.ADMIN and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        grading_period_id = request.data.get("grading_period_id")
        if not grading_period_id:
            return Response({"detail": "grading_period_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        submission = GradeSubmission.objects.filter(
            course_section=course_section,
            grading_period_id=grading_period_id,
        ).first()

        if not submission:
            return Response({"detail": "No submission found for this period."}, status=status.HTTP_404_NOT_FOUND)

        # Can only take back from submitted status (not published)
        if submission.status != GradeSubmissionStatus.SUBMITTED:
            return Response({"detail": f"Cannot take back: current status is {submission.status}."}, status=status.HTTP_400_BAD_REQUEST)

        submission.status = GradeSubmissionStatus.DRAFT
        submission.taken_back_at = timezone.now()
        submission.save()

        return Response({
            "submission_id": str(submission.id),
            "status": submission.status,
        })


class ReportCardPublishView(APIView):
    """Adviser publishes report card for a section and grading period."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, section_id):
        from core.models import TeacherAdvisory

        section = Section.objects.filter(id=section_id).first()
        if not section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Only the active adviser or admin can publish
        if request.user.role != User.Role.ADMIN:
            is_adviser = TeacherAdvisory.objects.filter(
                teacher=request.user,
                section=section,
                is_active=True,
            ).exists()
            if not is_adviser:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        grading_period_id = request.data.get("grading_period_id")
        if not grading_period_id:
            return Response({"detail": "grading_period_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        grading_period = GradingPeriod.objects.filter(id=grading_period_id).first()
        if not grading_period:
            return Response({"detail": "Grading period not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get or create the report card
        report_card, created = SectionReportCard.objects.get_or_create(
            section=section,
            grading_period=grading_period,
        )

        report_card.is_published = True
        report_card.published_by = request.user
        report_card.published_at = timezone.now()
        report_card.save()

        # Set all related GradeSubmissions to published
        course_sections = CourseSection.objects.filter(section=section, is_active=True)
        updated_count = GradeSubmission.objects.filter(
            course_section__in=course_sections,
            grading_period=grading_period,
            status=GradeSubmissionStatus.SUBMITTED,
        ).update(status=GradeSubmissionStatus.PUBLISHED)

        # Recompute Enrollment.final_grade for all students in this section for this period
        enrollments = Enrollment.objects.filter(
            course_section__in=course_sections,
            is_active=True,
        )
        for enrollment in enrollments:
            # Final grade = average of all published GradeEntry scores
            entries = GradeEntry.objects.filter(
                enrollment=enrollment,
                is_published=True,
            )
            scores = [e.score for e in entries if e.score is not None]
            if scores:
                enrollment.final_grade = round(sum(scores) / len(scores), 2)
            else:
                enrollment.final_grade = None
            enrollment.save(update_fields=["final_grade"])

        return Response({
            "report_card_id": str(report_card.id),
            "is_published": report_card.is_published,
            "published_at": report_card.published_at,
            "submissions_published_count": updated_count,
        })


class ReportCardUnpublishView(APIView):
    """Adviser unpublishes report card for a section and grading period."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, section_id):
        from core.models import TeacherAdvisory

        section = Section.objects.filter(id=section_id).first()
        if not section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role != User.Role.ADMIN:
            is_adviser = TeacherAdvisory.objects.filter(
                teacher=request.user,
                section=section,
                is_active=True,
            ).exists()
            if not is_adviser:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        grading_period_id = request.data.get("grading_period_id")
        if not grading_period_id:
            return Response({"detail": "grading_period_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        report_card = SectionReportCard.objects.filter(
            section=section,
            grading_period_id=grading_period_id,
        ).first()

        if not report_card:
            return Response({"detail": "No report card found for this period."}, status=status.HTTP_404_NOT_FOUND)

        if not report_card.is_published:
            return Response({"detail": "Report card is not published."}, status=status.HTTP_400_BAD_REQUEST)

        report_card.is_published = False
        report_card.save()

        # Revert all related GradeSubmissions from published to submitted
        course_sections = CourseSection.objects.filter(section=section, is_active=True)
        GradeSubmission.objects.filter(
            course_section__in=course_sections,
            grading_period_id=grading_period_id,
            status=GradeSubmissionStatus.PUBLISHED,
        ).update(status=GradeSubmissionStatus.SUBMITTED)

        return Response({
            "report_card_id": str(report_card.id),
            "is_published": report_card.is_published,
        })


class AdviserOverrideView(APIView):
    """Adviser directly overrides a grade entry."""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        # pk = grade_entry_id
        grade_entry = GradeEntry.objects.filter(id=pk).select_related(
            "enrollment__course_section__section"
        ).first()
        if not grade_entry:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Only the adviser of the section or admin can override
        if request.user.role != User.Role.ADMIN:
            from core.models import TeacherAdvisory
            is_adviser = TeacherAdvisory.objects.filter(
                teacher=request.user,
                section=grade_entry.enrollment.course_section.section,
                is_active=True,
            ).exists()
            if not is_adviser:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Only allowed when GradeSubmission.status = submitted (not published)
        submission = GradeSubmission.objects.filter(
            course_section=grade_entry.enrollment.course_section,
            grading_period=grade_entry.grading_period,
        ).first()

        if not submission or submission.status != GradeSubmissionStatus.SUBMITTED:
            return Response(
                {"detail": "Can only override grades when submission status is submitted."},
                status=status.HTTP_400_BAD_REQUEST
            )

        override_score = request.data.get("override_score")
        if override_score is None:
            return Response({"detail": "override_score is required."}, status=status.HTTP_400_BAD_REQUEST)

        previous_score = grade_entry.score
        new_score = Decimal(str(override_score))

        # Create audit log
        AdviserOverrideLog.objects.create(
            grade_entry=grade_entry,
            adviser=request.user,
            previous_score=previous_score if previous_score is not None else Decimal("0"),
            new_score=new_score,
        )

        grade_entry.override_score = new_score
        grade_entry.adviser_overridden = True
        grade_entry.save(update_fields=["override_score", "adviser_overridden"])

        return Response({
            "grade_entry_id": str(grade_entry.id),
            "previous_score": float(previous_score) if previous_score is not None else None,
            "new_score": float(new_score),
            "adviser_overridden": True,
        })


class StudentReportCardView(APIView):
    """Student's own report card — only shows published periods."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        enrollments = list(
            Enrollment.objects.filter(
                student=request.user, is_active=True, course_section__is_active=True
            ).select_related("course_section__course", "course_section__section", "course_section__teacher")
        )

        if not enrollments:
            return Response({"subjects": [], "periods": [], "overall_average": None})

        # Determine which sections the student is in
        sections = {e.course_section.section for e in enrollments}

        # Get all published report cards for these sections
        published_report_cards = SectionReportCard.objects.filter(
            section__in=sections,
            is_published=True,
        ).select_related("grading_period")

        published_period_ids = [rc.grading_period_id for rc in published_report_cards]

        if not published_period_ids:
            return Response({
                "section_name": None,
                "grade_level": None,
                "strand": None,
                "school_year": None,
                "periods": [],
                "subjects": [],
                "overall_average": None,
            })

        # Get grading period details
        grading_periods = GradingPeriod.objects.filter(id__in=published_period_ids).order_by("period_number")

        # Build subjects with period grades
        subjects = []
        all_final_grades = []

        for enrollment in enrollments:
            cs = enrollment.course_section
            course = cs.course
            teacher = cs.teacher

            # Get published grade entries for this enrollment
            entries = GradeEntry.objects.filter(
                enrollment=enrollment,
                grading_period_id__in=published_period_ids,
            ).select_related("grading_period")

            period_grades = []
            entry_scores = []

            for gp in grading_periods:
                entry = next((e for e in entries if e.grading_period_id == gp.id), None)
                score = float(entry.score) if entry and entry.score is not None else None
                period_grades.append({
                    "period_label": gp.label,
                    "score": score,
                    "adviser_overridden": entry.adviser_overridden if entry else False,
                })
                if score is not None:
                    entry_scores.append(score)

            # Final grade for this subject = average of period scores
            final_grade = round(sum(entry_scores) / len(entry_scores), 2) if entry_scores else None
            if final_grade is not None:
                all_final_grades.append(final_grade)

            subjects.append({
                "course_section_id": str(cs.id),
                "course_code": course.code,
                "course_title": course.title,
                "teacher_name": teacher.full_name if teacher else None,
                "period_grades": period_grades,
                "final_grade": final_grade,
                "final_grade_letter": _letter_grade(Decimal(str(final_grade))) if final_grade is not None else None,
            })

        overall_average = round(sum(all_final_grades) / len(all_final_grades), 2) if all_final_grades else None

        # Get section info from the first enrollment
        first_section = list(sections)[0] if sections else None

        return Response({
            "section_name": first_section.name if first_section else None,
            "grade_level": first_section.grade_level if first_section else None,
            "strand": first_section.strand if first_section else None,
            "school_year": enrollments[0].course_section.school_year if enrollments else None,
            "periods": [
                {"id": str(gp.id), "label": gp.label, "period_number": gp.period_number}
                for gp in grading_periods
            ],
            "subjects": subjects,
            "overall_average": overall_average,
        })


__all__ = [
    'CourseSectionGradesView',
    'CourseSectionGradebookView',
    'CourseSectionGradesExportCSVView',
    'EnrollmentGradeOverrideView',
    'GradingPeriodListView',
    'StudentGradesView',
    'AdvisoryGradesView',
    'SubjectGradesView',
    'GradeEntryUpdateView',
    'GradeEntryPublishView',
    'BulkPublishGradesView',
    'ComputeFinalGradeView',
    'BulkPublishFinalGradesView',
    'GradeWeightConfigView',
    'GradeSubmissionSubmitView',
    'GradeSubmissionTakeBackView',
    'ReportCardPublishView',
    'ReportCardUnpublishView',
    'AdviserOverrideView',
    'StudentReportCardView',
]
