import csv
import logging
import os
import shutil
import subprocess
from django.contrib.auth import authenticate
from django.core.cache import cache
from django.core.files.base import File
from django.db.models import Avg, Count, Q, Sum
from django.db.utils import OperationalError
from django.core.files.storage import default_storage
from django.http import StreamingHttpResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta
from io import StringIO, BytesIO
from decimal import Decimal, ROUND_HALF_UP
from rest_framework import mixins, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    Activity,
    ActivityComment,
    ActivityReminder,
    Announcement,
    AssignmentGroup,
    AttendanceRecord,
    CalendarEvent,
    Course,
    CourseFile,
    CourseSection,
    Enrollment,
    MeetingSession,
    Notification,
    PasswordResetRequest,
    PushToken,
    Quiz,
    QuizAnswer,
    QuizAttempt,
    QuizChoice,
    QuizQuestion,
    Section,
    Submission,
    TodoItem,
    User,
    WeeklyModule,
)
from .serializers import (
    ActivityCommentSerializer,
    ActivitySerializer,
    ActivityReminderSerializer,
    AssignmentGroupSerializer,
    AnnouncementSerializer,
    AttendanceRecordSerializer,
    CalendarEventSerializer,
    CourseFileSerializer,
    MeetingSessionSerializer,
    NotificationSerializer,
    PasswordResetRequestSerializer,
    PushTokenSerializer,
    QuizAnswerGradeSerializer,
    QuizAnswerInputSerializer,
    QuizQuestionWriteSerializer,
    QuizQuestionStudentSerializer,
    QuizSerializer,
    SubmissionGradeSerializer,
    SubmissionSerializer,
    TodoItemSerializer,
    UserCreateSerializer,
    UserSerializer,
    WeeklyModuleSerializer,
)
from .permissions import IsAdminRole
from .serializers import (
    ActivityCommentSerializer,
    ActivitySerializer,
    ActivityReminderSerializer,
    AssignmentGroupSerializer,
    AnnouncementSerializer,
    AttendanceRecordSerializer,
    CalendarEventSerializer,
    CourseFileSerializer,
    MeetingSessionSerializer,
    NotificationSerializer,
    PasswordResetRequestSerializer,
    PushTokenSerializer,
    QuizAnswerGradeSerializer,
    QuizAnswerInputSerializer,
    QuizQuestionWriteSerializer,
    QuizQuestionStudentSerializer,
    QuizSerializer,
    SubmissionGradeSerializer,
    SubmissionSerializer,
    TodoItemSerializer,
    UserCreateSerializer,
    UserSerializer,
    WeeklyModuleSerializer,
)
from .push_notifications import send_push_notification_to_users


def _notify_students_for_course_section(
    *,
    course_section: CourseSection,
    notif_type: str,
    title: str,
    body: str,
    activity: Activity | None = None,
    quiz: Quiz | None = None,
    throttle_seconds: int = 90,
):
    student_ids = list(
        Enrollment.objects.filter(
            course_section=course_section,
            is_active=True,
            student__status=User.Status.ACTIVE,
        )
        .values_list("student_id", flat=True)
        .distinct()
    )
    if not student_ids:
        return

    recent_cutoff = timezone.now() - timedelta(seconds=max(int(throttle_seconds), 0))
    existing_recipient_ids = set(
        Notification.objects.filter(
            recipient_id__in=student_ids,
            type=notif_type,
            title=title,
            course_section=course_section,
            activity=activity,
            quiz=quiz,
            created_at__gte=recent_cutoff,
        ).values_list("recipient_id", flat=True)
    )

    rows = []
    for student_id in student_ids:
        if student_id in existing_recipient_ids:
            continue
        rows.append(
            Notification(
                recipient_id=student_id,
                type=notif_type,
                title=title,
                body=body,
                course_section=course_section,
                activity=activity,
                quiz=quiz,
            )
        )
    if not rows:
        return
    Notification.objects.bulk_create(rows)

    # Send push notifications to students
    # Get student IDs who didn't already receive this notification
    new_recipient_ids = [str(student_id) for student_id in student_ids if student_id not in existing_recipient_ids]

    if new_recipient_ids:
        # Build data payload for deep linking
        data = {
            "type": notif_type,
            "course_section_id": str(course_section.id),
        }
        if activity:
            data["activity_id"] = str(activity.id)
        if quiz:
            data["quiz_id"] = str(quiz.id)

        # Send push notifications asynchronously (in production, use Celery task)
        try:
            send_push_notification_to_users(
                user_ids=new_recipient_ids,
                title=title,
                body=body,
                data=data,
            )
        except Exception as e:
            logger.warning(f"Failed to send push notifications: {e}")


LETTER_SCALE = [
    (Decimal("97"), "A+"),
    (Decimal("93"), "A"),
    (Decimal("90"), "A-"),
    (Decimal("87"), "B+"),
    (Decimal("83"), "B"),
    (Decimal("80"), "B-"),
    (Decimal("77"), "C+"),
    (Decimal("73"), "C"),
    (Decimal("70"), "C-"),
    (Decimal("60"), "D"),
    (Decimal("0"), "F"),
]

logger = logging.getLogger(__name__)
OFFICE_CONVERTIBLE_EXTENSIONS = {"doc", "docx", "ppt", "pptx", "xls", "xlsx"}
GRADE_COMPONENT_WEIGHTS = {
    "activities": Decimal("70"),
    "quizzes": Decimal("20"),
    "attendance": Decimal("10"),
}


def _letter_grade(value: Decimal | None) -> str | None:
    if value is None:
        return None
    for min_score, letter in LETTER_SCALE:
        if value >= min_score:
            return letter
    return "F"


def _quantize_pct(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _compute_activity_percentage(enrollment: Enrollment) -> Decimal | None:
    activities = list(
        Activity.objects.filter(
            course_section=enrollment.course_section,
            is_published=True,
            points__gt=0,
        ).select_related("assignment_group")
    )
    if not activities:
        return None

    # Get all submissions for the student for these activities
    all_submissions = Submission.objects.filter(activity__in=activities, student=enrollment.student)
    submissions_by_activity: dict[str, list[Submission]] = {}
    for s in all_submissions:
        activity_id = str(s.activity_id)
        if activity_id not in submissions_by_activity:
            submissions_by_activity[activity_id] = []
        submissions_by_activity[activity_id].append(s)

    grouped: dict[str, dict[str, Decimal]] = {}
    explicit_weights: dict[str, Decimal] = {}
    explicit_weight_sum = Decimal("0")

    for activity in activities:
        group_key = str(activity.assignment_group_id) if activity.assignment_group_id else "__default__"
        possible = Decimal(activity.points or 0)

        # Get all submissions for this activity
        activity_submissions = submissions_by_activity.get(str(activity.id), [])
        earned = Decimal("0")

        if activity_submissions:
            # Calculate best score based on policy
            scores = [s.score for s in activity_submissions if s.score is not None]
            if scores:
                if activity.score_selection_policy == Activity.ScorePolicy.HIGHEST:
                    earned = max(scores)
                else:  # LATEST
                    # Get the latest submission's score
                    latest = max(activity_submissions, key=lambda s: s.attempt_number)
                    earned = Decimal(latest.score) if latest.score is not None else Decimal("0")

        data = grouped.setdefault(group_key, {"earned": Decimal("0"), "possible": Decimal("0")})
        data["earned"] += earned
        data["possible"] += possible

        if activity.assignment_group and activity.assignment_group.weight_percent is not None:
            weight = Decimal(activity.assignment_group.weight_percent)
            explicit_weights[group_key] = weight

    explicit_weight_sum = sum(explicit_weights.values(), Decimal("0"))
    weighted_enabled = explicit_weight_sum > Decimal("0")

    if weighted_enabled:
        default_weight = max(Decimal("0"), Decimal("100") - explicit_weight_sum)
        raw_weights: dict[str, Decimal] = {}
        for key in grouped.keys():
            raw_weights[key] = explicit_weights.get(key, default_weight if key == "__default__" else Decimal("0"))

        total_weight = sum(raw_weights.values(), Decimal("0"))
        if total_weight <= Decimal("0"):
            weighted_enabled = False
        else:
            normalized = {k: (w * Decimal("100") / total_weight) for k, w in raw_weights.items()}
            weighted_total = Decimal("0")
            for key, data in grouped.items():
                possible = data["possible"]
                if possible <= 0:
                    continue
                group_pct = (data["earned"] / possible) * Decimal("100")
                weighted_total += (group_pct * normalized.get(key, Decimal("0")) / Decimal("100"))
            return _quantize_pct(max(Decimal("0"), min(weighted_total, Decimal("100"))))

    total_possible = sum(d["possible"] for d in grouped.values())
    if total_possible <= 0:
        return None
    total_earned = sum(d["earned"] for d in grouped.values())
    pct = (total_earned / total_possible) * Decimal("100")
    return _quantize_pct(max(Decimal("0"), min(pct, Decimal("100"))))


def _compute_quiz_percentage(enrollment: Enrollment) -> Decimal | None:
    quizzes = list(
        Quiz.objects.filter(course_section=enrollment.course_section, is_published=True).order_by("created_at", "id")
    )
    if not quizzes:
        return None

    quiz_totals_raw = (
        QuizQuestion.objects.filter(quiz__in=quizzes)
        .values("quiz_id")
        .annotate(total=Sum("points"))
    )
    quiz_possible_map = {
        str(row["quiz_id"]): Decimal(str(row["total"] or 0))
        for row in quiz_totals_raw
    }

    # Get all attempts for the student
    attempts = QuizAttempt.objects.filter(
        quiz__in=quizzes,
        student=enrollment.student,
        is_submitted=True,
        score__isnull=False,
    ).order_by("quiz_id", "attempt_number")

    # Group attempts by quiz and calculate score based on policy
    attempts_by_quiz: dict[str, list[QuizAttempt]] = {}
    for attempt in attempts:
        quiz_id = str(attempt.quiz_id)
        if quiz_id not in attempts_by_quiz:
            attempts_by_quiz[quiz_id] = []
        attempts_by_quiz[quiz_id].append(attempt)

    # Build a map of quiz -> selected score based on policy
    score_by_quiz: dict[str, Decimal] = {}
    quiz_policy_map = {str(q.id): q.score_selection_policy for q in quizzes}

    for quiz_id, quiz_attempts in attempts_by_quiz.items():
        scores = [Decimal(str(a.score or 0)) for a in quiz_attempts if a.score is not None]
        if not scores:
            continue
        policy = quiz_policy_map.get(quiz_id, Quiz.ScorePolicy.HIGHEST)
        if policy == Quiz.ScorePolicy.HIGHEST:
            score_by_quiz[quiz_id] = max(scores)
        else:  # LATEST - use the last attempt's score
            score_by_quiz[quiz_id] = scores[-1]

    total_possible = Decimal("0")
    total_earned = Decimal("0")
    for quiz in quizzes:
        possible = quiz_possible_map.get(str(quiz.id), Decimal("0"))
        if possible <= 0:
            continue
        total_possible += possible
        earned = score_by_quiz.get(str(quiz.id), Decimal("0"))
        total_earned += max(Decimal("0"), min(earned, possible))

    if total_possible <= 0:
        return None
    pct = (total_earned / total_possible) * Decimal("100")
    return _quantize_pct(max(Decimal("0"), min(pct, Decimal("100"))))


def _compute_attendance_percentage(enrollment: Enrollment) -> Decimal | None:
    total_sessions = MeetingSession.objects.filter(course_section=enrollment.course_section).count()
    if total_sessions <= 0:
        return None

    status_counts = {
        row["status"]: row["count"]
        for row in (
            AttendanceRecord.objects.filter(
                meeting__course_section=enrollment.course_section,
                student=enrollment.student,
            )
            .values("status")
            .annotate(count=Count("id"))
        )
    }
    present = Decimal(str(status_counts.get(AttendanceRecord.AttendanceStatus.PRESENT, 0)))
    excused = Decimal(str(status_counts.get(AttendanceRecord.AttendanceStatus.EXCUSED, 0)))
    late = Decimal(str(status_counts.get(AttendanceRecord.AttendanceStatus.LATE, 0)))
    attended_units = present + excused + (late * Decimal("0.5"))
    pct = (attended_units / Decimal(str(total_sessions))) * Decimal("100")
    return _quantize_pct(max(Decimal("0"), min(pct, Decimal("100"))))


def _compute_enrollment_grade(enrollment: Enrollment) -> Decimal | None:
    if enrollment.manual_final_grade is not None:
        return _quantize_pct(Decimal(enrollment.manual_final_grade))

    components = {
        "activities": _compute_activity_percentage(enrollment),
        "quizzes": _compute_quiz_percentage(enrollment),
        "attendance": _compute_attendance_percentage(enrollment),
    }
    active_components = [(name, pct) for name, pct in components.items() if pct is not None]
    if not active_components:
        return None

    active_weight_total = sum((GRADE_COMPONENT_WEIGHTS.get(name, Decimal("0")) for name, _ in active_components), Decimal("0"))
    if active_weight_total <= 0:
        return None

    total = Decimal("0")
    for name, pct in active_components:
        weight = GRADE_COMPONENT_WEIGHTS.get(name, Decimal("0"))
        total += (pct * weight) / active_weight_total
    return _quantize_pct(max(Decimal("0"), min(total, Decimal("100"))))


def _recompute_enrollment_grade(enrollment: Enrollment) -> Enrollment:
    enrollment.final_grade = _compute_enrollment_grade(enrollment)
    enrollment.save(update_fields=["final_grade"])
    return enrollment


def _get_grade_summary_metadata(enrollment: Enrollment) -> dict:
    """
    Compute grade summary metadata for a student's enrollment.
    Returns counts of graded/pending/excluded items for frontend badge display.
    """
    from datetime import datetime

    # Get activities and quizzes for this course section
    activities = list(
        Activity.objects.filter(
            course_section=enrollment.course_section,
            is_published=True,
            points__gt=0,
        )
    )
    quizzes = list(
        Quiz.objects.filter(
            course_section=enrollment.course_section,
            is_published=True,
        )
    )

    enrolled_at = enrollment.enrolled_at

    # Count activities
    graded_activities = 0
    pending_activities = 0
    excluded_activities = 0
    total_activities = len(activities)

    for activity in activities:
        # Check if activity deadline is before enrollment (pre-enrollment exclusion)
        if activity.deadline and enrolled_at:
            activity_deadline = activity.deadline
            if isinstance(activity_deadline, str):
                activity_deadline = datetime.fromisoformat(activity_deadline.replace('Z', '+00:00'))
            enrolled_at_dt = enrolled_at if not isinstance(enrolled_at, str) else datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
            if activity_deadline < enrolled_at_dt:
                excluded_activities += 1
                continue

        # Check submission status
        submission = Submission.objects.filter(
            activity=activity,
            student=enrollment.student
        ).order_by('-attempt_number').first()

        if submission:
            if submission.score is not None:
                graded_activities += 1
            else:
                # Submitted but not graded yet
                pending_activities += 1

    # Count quizzes
    graded_quizzes = 0
    pending_quizzes = 0
    excluded_quizzes = 0
    total_quizzes = len(quizzes)

    for quiz in quizzes:
        # Check if quiz close_at is before enrollment (pre-enrollment exclusion)
        if quiz.close_at and enrolled_at:
            close_at = quiz.close_at
            if isinstance(close_at, str):
                close_at = datetime.fromisoformat(close_at.replace('Z', '+00:00'))
            enrolled_at_dt = enrolled_at if not isinstance(enrolled_at, str) else datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
            if close_at < enrolled_at_dt:
                excluded_quizzes += 1
                continue

        # Check quiz attempt
        attempts = QuizAttempt.objects.filter(
            quiz=quiz,
            student=enrollment.student,
            is_submitted=True,
        )
        if attempts.exists():
            # Check if any attempt has a score
            has_score = attempts.filter(score__isnull=False).exists()
            if has_score:
                graded_quizzes += 1
            else:
                pending_quizzes += 1

    graded_items_count = graded_activities + graded_quizzes
    pending_items_count = pending_activities + pending_quizzes
    excluded_items_count = excluded_activities + excluded_quizzes
    total_items_count = (total_activities - excluded_activities) + (total_quizzes - excluded_quizzes)

    # Determine grade status
    has_pending = pending_items_count > 0
    has_released_grades = graded_items_count > 0
    has_no_gradeable_items = total_items_count == 0

    return {
        "graded_items_count": graded_items_count,
        "total_items_count": total_items_count,
        "pending_items_count": pending_items_count,
        "excluded_items_count": excluded_items_count,
        "has_pending": has_pending,
        "has_released_grades": has_released_grades,
        "has_no_gradeable_items": has_no_gradeable_items,
        "is_partial": has_pending and has_released_grades,
        "enrolled_at": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None,
    }


def _recompute_course_section_grades(course_section: CourseSection):
    enrollments = Enrollment.objects.filter(course_section=course_section, is_active=True).select_related("student")
    for enrollment in enrollments:
        _recompute_enrollment_grade(enrollment)


def _get_grade_summary_metadata(enrollment: Enrollment) -> dict:
    """
    Compute grade summary metadata for a student's enrollment.
    Returns counts of graded/pending/total items without modifying grade computation.
    """
    from datetime import datetime
    from decimal import Decimal

    course_section = enrollment.course_section
    enrolled_at = enrollment.enrolled_at

    # Get all published activities and quizzes
    activities = Activity.objects.filter(
        course_section=course_section,
        is_published=True,
        points__gt=0,
    ).only('id', 'deadline', 'points')

    quizzes = Quiz.objects.filter(
        course_section=course_section,
        is_published=True,
    ).only('id', 'close_at')

    # Get student's submissions and attempts
    activity_ids = [a.id for a in activities]
    quiz_ids = [q.id for q in quizzes]

    submissions = {
        str(s.activity_id): s
        for s in Submission.objects.filter(
            activity_id__in=activity_ids,
            student=enrollment.student
        )
    }

    attempts = list(QuizAttempt.objects.filter(
        quiz_id__in=quiz_ids,
        student=enrollment.student,
        is_submitted=True,
    ))

    attempts_by_quiz: dict[str, list] = {}
    for a in attempts:
        key = str(a.quiz_id)
        if key not in attempts_by_quiz:
            attempts_by_quiz[key] = []
        attempts_by_quiz[key].append(a)

    graded_items_count = 0
    total_items_count = 0
    pending_count = 0
    excluded_count = 0

    # Process activities
    for activity in activities:
        total_items_count += 1

        # Check if activity deadline is before enrollment (pre-enrollment exclusion)
        if activity.deadline and enrolled_at:
            activity_deadline = activity.deadline
            if isinstance(enrolled_at, str):
                enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
            else:
                enrolled_at_dt = enrolled_at
            if isinstance(activity_deadline, str):
                activity_deadline_dt = datetime.fromisoformat(activity_deadline.replace('Z', '+00:00'))
            else:
                activity_deadline_dt = activity_deadline

            if activity_deadline_dt < enrolled_at_dt:
                excluded_count += 1
                continue

        sub = submissions.get(str(activity.id))
        if sub:
            if sub.score is not None:
                graded_items_count += 1
            elif sub.status in (Submission.SubmissionStatus.SUBMITTED, Submission.SubmissionStatus.LATE):
                pending_count += 1

    # Process quizzes
    for quiz in quizzes:
        # Quizzes don't have pre-enrollment exclusion in the same way
        # but we check close_at similar to activities
        total_items_count += 1

        if quiz.close_at and enrolled_at:
            close_at = quiz.close_at
            if isinstance(enrolled_at, str):
                enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
            else:
                enrolled_at_dt = enrolled_at
            if isinstance(close_at, str):
                close_at_dt = datetime.fromisoformat(close_at.replace('Z', '+00:00'))
            else:
                close_at_dt = close_at

            if close_at_dt < enrolled_at_dt:
                excluded_count += 1
                continue

        quiz_attempts = attempts_by_quiz.get(str(quiz.id), [])
        if quiz_attempts:
            # Check if any attempt has a score
            has_score = any(a.score is not None for a in quiz_attempts)
            if has_score:
                graded_items_count += 1
            else:
                pending_count += 1

    return {
        "graded_items_count": graded_items_count,
        "total_items_count": total_items_count,
        "pending_count": pending_count,
        "excluded_count": excluded_count,
        "has_pending": pending_count > 0,
        "has_grades": graded_items_count > 0,
        "has_items": total_items_count > 0,
    }


def _sync_student_activity_items(student: User):
    if student.role != User.Role.STUDENT:
        return

    activities = list(
        Activity.objects.filter(
            is_published=True,
            course_section__enrollments__student=student,
            course_section__enrollments__is_active=True,
        )
        .select_related("course_section")
        .distinct()
    )
    quizzes = list(
        Quiz.objects.filter(
            is_published=True,
            course_section__enrollments__student=student,
            course_section__enrollments__is_active=True,
        )
        .select_related("course_section")
        .distinct()
    )
    activity_ids = [a.id for a in activities]
    quiz_ids = [q.id for q in quizzes]
    submissions = {
        s.activity_id: s
        for s in Submission.objects.filter(student=student, activity_id__in=activity_ids)
    }
    quiz_attempts = (
        QuizAttempt.objects.filter(student=student, quiz_id__in=quiz_ids, is_submitted=True)
        .order_by("quiz_id", "-attempt_number", "-submitted_at")
    )
    latest_attempt_by_quiz: dict[str, QuizAttempt] = {}
    for attempt in quiz_attempts:
        key = str(attempt.quiz_id)
        if key not in latest_attempt_by_quiz:
            latest_attempt_by_quiz[key] = attempt

    for activity in activities:
        sub = submissions.get(activity.id)
        is_done = bool(sub and sub.submitted_at)
        TodoItem.objects.update_or_create(
            user=student,
            activity=activity,
            defaults={
                "title": f"{activity.title}",
                "description": activity.instructions or activity.description or "",
                "due_at": activity.deadline,
                "is_done": is_done,
                "completed_at": sub.submitted_at if is_done else None,
            },
        )
        CalendarEvent.objects.update_or_create(
            creator=student,
            activity=activity,
            defaults={
                "course_section": activity.course_section,
                "title": activity.title,
                "description": activity.instructions or activity.description or "",
                "event_type": CalendarEvent.EventType.DEADLINE,
                "start_at": activity.deadline or timezone.now(),
                "end_at": activity.deadline or None,
                "all_day": activity.deadline is None,
                "is_personal": False,  # Activity-generated events are not user-created
                "color": None,
            },
        )

    for quiz in quizzes:
        attempt = latest_attempt_by_quiz.get(str(quiz.id))
        is_done = bool(attempt and attempt.submitted_at)
        TodoItem.objects.update_or_create(
            user=student,
            quiz=quiz,
            defaults={
                "title": f"{quiz.title}",
                "description": quiz.instructions or "",
                "due_at": quiz.close_at,
                "is_done": is_done,
                "completed_at": attempt.submitted_at if is_done else None,
                "activity": None,
            },
        )

    TodoItem.objects.filter(user=student, activity__isnull=False).exclude(activity_id__in=activity_ids).delete()
    TodoItem.objects.filter(user=student, quiz__isnull=False).exclude(quiz_id__in=quiz_ids).delete()
    CalendarEvent.objects.filter(creator=student, activity__isnull=False).exclude(activity_id__in=activity_ids).delete()


def _sync_course_section_students_activity_items(course_section: CourseSection):
    student_ids = Enrollment.objects.filter(
        course_section=course_section,
        is_active=True,
        student__role=User.Role.STUDENT,
    ).values_list("student_id", flat=True)
    for student in User.objects.filter(id__in=student_ids):
        _sync_student_activity_items(student)


def _sync_student_items_best_effort(student: User, *, min_interval_seconds: int = 12):
    if student.role != User.Role.STUDENT:
        return
    cooldown_key = f"sync-student-items:cooldown:{student.id}"
    lock_key = f"sync-student-items:lock:{student.id}"
    if cache.get(cooldown_key):
        return
    if not cache.add(lock_key, "1", timeout=8):
        return
    try:
        _sync_student_activity_items(student)
        cache.set(cooldown_key, "1", timeout=max(1, int(min_interval_seconds)))
    except OperationalError:
        logger.warning("todo_sync_skipped_db_locked student_id=%s", student.id)
    finally:
        cache.delete(lock_key)


def _build_storage_abs_path(path: str) -> str | None:
    if hasattr(default_storage, "path"):
        try:
            return default_storage.path(path)
        except Exception:
            return None
    return None


def _convert_office_upload_to_pdf_preview(*, request, source_storage_path: str, original_name: str) -> str | None:
    ext = (os.path.splitext(original_name)[1] or "").lower().replace(".", "")
    if ext not in OFFICE_CONVERTIBLE_EXTENSIONS:
        return None

    source_abs = _build_storage_abs_path(source_storage_path)
    if not source_abs or not os.path.exists(source_abs):
        return None

    libreoffice_bin = shutil.which("libreoffice")
    if not libreoffice_bin:
        logger.warning("preview_conversion_skipped_no_libreoffice source=%s", source_storage_path)
        return None

    output_dir = os.path.join(os.path.dirname(source_abs), "__previews__")
    os.makedirs(output_dir, exist_ok=True)

    try:
        subprocess.run(
            [libreoffice_bin, "--headless", "--convert-to", "pdf", "--outdir", output_dir, source_abs],
            check=False,
            timeout=90,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception:
        logger.exception("preview_conversion_failed source=%s", source_storage_path)
        return None

    generated_pdf_abs = os.path.join(output_dir, f"{os.path.splitext(os.path.basename(source_abs))[0]}.pdf")
    if not os.path.exists(generated_pdf_abs):
        logger.warning("preview_conversion_no_output source=%s", source_storage_path)
        return None

    base_name = os.path.splitext(os.path.basename(original_name))[0].replace(" ", "_")
    preview_rel = f"course_files/previews/{request.user.id}/{timezone.now().timestamp()}_{base_name}.pdf"
    try:
        with open(generated_pdf_abs, "rb") as fp:
            saved_preview = default_storage.save(preview_rel, File(fp))
        return request.build_absolute_uri(default_storage.url(saved_preview))
    except Exception:
        logger.exception("preview_storage_failed source=%s", source_storage_path)
        return None


class AuthLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")
        if not email or not password:
            return Response({"detail": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username=email, password=password)
        if not user:
            return Response({"detail": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)

        if user.status != User.Status.ACTIVE:
            return Response({"detail": "This account is inactive."}, status=status.HTTP_403_FORBIDDEN)

        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": UserSerializer(user, context={"request": request}).data,
            }
        )


class MeView(APIView):
    def get(self, request):
        return Response(UserSerializer(request.user, context={"request": request}).data)


class ChangePasswordView(APIView):
    """Change password for authenticated users. Used for first-time setup."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        new_password = request.data.get("new_password")

        if not new_password:
            return Response(
                {"detail": "New password is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if len(new_password) < 8:
            return Response(
                {"detail": "Password must be at least 8 characters long."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set the new password (this hashes it)
        user.set_password(new_password)
        user.requires_setup = False
        # Save all fields to ensure password is properly persisted
        user.save()

        # Generate new tokens for the user
        refresh = RefreshToken.for_user(user)

        return Response({
            "detail": "Password changed successfully.",
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": UserSerializer(user, context={"request": request}).data,
        })


class ForgotPasswordRequestView(APIView):
    """Request a password reset. Creates a pending request for admin approval."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        from django.utils import timezone as tz
        from .email_utils import generate_random_password, send_password_reset_email

        email = request.data.get("email")
        if not email:
            return Response(
                {"detail": "Email is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Find user by school email
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            # Don't reveal if user exists or not
            return Response({
                "detail": "If an account with that email exists, a password reset request has been submitted. Please wait for admin approval."
            })

        # Only teachers and students can request password reset
        if user.role not in [User.Role.TEACHER, User.Role.STUDENT]:
            return Response(
                {"detail": "Only teachers and students can request password reset."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Check if user has personal email
        if not user.personal_email:
            return Response(
                {"detail": "No personal email configured for this account. Please contact administrator."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Rate limit: 1 request per hour per email
        cache_key = f"password_reset_rate:{email}"
        if cache.get(cache_key):
            return Response(
                {"detail": "A password reset request was recently submitted. Please wait before requesting another."},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        # Check for existing pending request
        existing_pending = PasswordResetRequest.objects.filter(
            user=user,
            status=PasswordResetRequest.Status.PENDING
        ).exists()

        if existing_pending:
            return Response({
                "detail": "A password reset request is already pending. Please wait for admin approval."
            })

        # Create the request
        PasswordResetRequest.objects.create(
            user=user,
            personal_email=user.personal_email,
        )

        # Set rate limit cache (1 hour)
        cache.set(cache_key, True, timeout=3600)

        return Response({
            "detail": "Password reset request submitted. You will receive an email once approved by administrator."
        })


class PasswordResetRequestViewSet(viewsets.ReadOnlyModelViewSet):
    """Admin viewset for managing password reset requests."""
    permission_classes = [IsAdminRole]
    queryset = PasswordResetRequest.objects.select_related("user", "resolved_by").all()
    serializer_class = PasswordResetRequestSerializer

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        """Approve a password reset request and send new credentials."""
        from .email_utils import generate_random_password, send_password_reset_email

        reset_request = self.get_object()

        if reset_request.status != PasswordResetRequest.Status.PENDING:
            return Response(
                {"detail": "This request has already been processed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        user = reset_request.user

        # Generate new password
        new_password = generate_random_password()
        user.set_password(new_password)
        user.requires_setup = True
        user.save(update_fields=["password", "requires_setup", "updated_at"])

        # Send email
        success, message = send_password_reset_email(user, new_password)

        if success:
            # Update request status
            reset_request.status = PasswordResetRequest.Status.APPROVED
            reset_request.resolved_at = timezone.now()
            reset_request.resolved_by = request.user
            reset_request.save()

            return Response({
                "detail": f"Password reset approved. New credentials sent to {user.personal_email}"
            })
        else:
            return Response(
                {"detail": f"Failed to send email: {message}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=["post"])
    def decline(self, request, pk=None):
        """Decline a password reset request."""
        reset_request = self.get_object()

        if reset_request.status != PasswordResetRequest.Status.PENDING:
            return Response(
                {"detail": "This request has already been processed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        reset_request.status = PasswordResetRequest.Status.DECLINED
        reset_request.resolved_at = timezone.now()
        reset_request.resolved_by = request.user
        reset_request.save()

        return Response({"detail": "Password reset request declined."})


class DashboardStatsView(APIView):
    permission_classes = [IsAdminRole]

    def get(self, request):
        data = {
            "students": User.objects.filter(role=User.Role.STUDENT).count(),
            "teachers": User.objects.filter(role=User.Role.TEACHER).count(),
            "courses": Course.objects.filter(is_active=True).count(),
            "sections": Section.objects.filter(is_active=True).count(),
        }
        return Response(data)


class ProfileViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by("last_name", "first_name")

    def get_permissions(self):
        if self.action in ["list", "create", "destroy", "toggle_status"]:
            return [IsAdminRole()]
        return [permissions.IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == "create":
            return UserCreateSerializer
        return UserSerializer

    def get_queryset(self):
        user = self.request.user
        role = self.request.query_params.get("role")

        qs = User.objects.all()
        if role:
            qs = qs.filter(role=role)

        if user.role != User.Role.ADMIN:
            return qs.filter(id=user.id)
        return qs.order_by("last_name", "first_name")

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop("partial", False)

        if request.user.role != User.Role.ADMIN and request.user.id != instance.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        serializer = UserSerializer(instance, data=request.data, partial=partial, context={"request": request})
        serializer.is_valid(raise_exception=True)

        if request.user.role != User.Role.ADMIN:
            allowed_fields = {"first_name", "last_name", "middle_name", "avatar_url", "theme", "section", "grade_level", "strand"}
            for key in list(serializer.validated_data.keys()):
                if key not in allowed_fields:
                    serializer.validated_data.pop(key, None)

        serializer.save()
        return Response(serializer.data)

    @action(detail=True, methods=["post"], permission_classes=[IsAdminRole])
    def toggle_status(self, request, pk=None):
        user = self.get_object()
        user.status = User.Status.INACTIVE if user.status == User.Status.ACTIVE else User.Status.ACTIVE
        user.save(update_fields=["status", "updated_at"])
        return Response(UserSerializer(user, context={"request": request}).data)


class AvatarUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "file is required."}, status=status.HTTP_400_BAD_REQUEST)

        request.user.avatar = file_obj
        request.user.save()

        # Return the updated user with avatar_url
        return Response(UserSerializer(request.user, context={"request": request}).data)


class TodoItemViewSet(viewsets.ModelViewSet):
    serializer_class = TodoItemSerializer

    def get_queryset(self):
        if self.request.user.role == User.Role.STUDENT:
            _sync_student_items_best_effort(self.request.user)
        return (
            TodoItem.objects.filter(user=self.request.user)
            .select_related("activity__course_section", "quiz__course_section")
            .order_by("is_done", "due_at")
        )

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class CalendarEventViewSet(viewsets.ModelViewSet):
    serializer_class = CalendarEventSerializer

    def get_queryset(self):
        if self.request.user.role == User.Role.STUDENT:
            _sync_student_items_best_effort(self.request.user)
        qs = CalendarEvent.objects.filter(Q(creator=self.request.user) | Q(is_personal=False))
        start = self.request.query_params.get("start")
        end = self.request.query_params.get("end")
        if start:
            qs = qs.filter(start_at__gte=start)
        if end:
            qs = qs.filter(start_at__lte=end)
        return qs.order_by("start_at")

    def perform_create(self, serializer):
        serializer.save(creator=self.request.user)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.creator_id != request.user.id and request.user.role != User.Role.ADMIN:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class NotificationViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = NotificationSerializer

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user).order_by("-created_at")[:50]

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save(update_fields=["is_read"])
        return Response(NotificationSerializer(notif).data)

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
        return Response({"ok": True})


class StudentCoursesView(APIView):
    def get(self, request):
        enrollments = (
            Enrollment.objects.filter(student=request.user, is_active=True, course_section__is_active=True)
            .select_related("course_section__course", "course_section__section", "course_section__teacher")
            .order_by("course_section__course__title")
        )
        data = []
        for e in enrollments:
            _recompute_enrollment_grade(e)
            cs = e.course_section
            course = cs.course
            sec = cs.section
            course_tag = f"{course.code}@{sec.strand}-{sec.name}" if sec.strand and sec.strand != "NONE" else f"{course.code}@{sec.name}"
            final_grade = float(e.final_grade) if e.final_grade is not None else None

            # Get grade metadata for badge display
            grade_metadata = _get_grade_summary_metadata(e)

            data.append(
                {
                    "student_id": str(request.user.id),
                    "course_section_id": str(cs.id),
                    "course_id": str(course.id),
                    "course_code": course.code,
                    "course_title": course.title,
                    "cover_image_url": course.cover_image_url,
                    "color_overlay": course.color_overlay,
                    "section_name": sec.name,
                    "strand": sec.strand,
                    "grade_level": sec.grade_level,
                    "final_grade": final_grade,
                    "final_grade_letter": _letter_grade(Decimal(str(final_grade))) if final_grade is not None else None,
                    "grade_overridden": e.manual_final_grade is not None,
                    "teacher_name": cs.teacher.full_name if cs.teacher else None,
                    "course_tag": course_tag,
                    "semester": cs.semester,
                    "school_year": cs.school_year,
                    # Grade metadata for badge display
                    "grade_summary": grade_metadata,
                }
            )
        return Response(data)


class TeacherCoursesView(APIView):
    def get(self, request):
        course_sections = (
            CourseSection.objects.filter(teacher=request.user, is_active=True)
            .select_related("course", "section")
            .annotate(student_count=Count("enrollments", filter=Q(enrollments__is_active=True)))
            .order_by("course__title")
        )
        data = []
        for cs in course_sections:
            course = cs.course
            sec = cs.section
            course_tag = f"{course.code}@{sec.strand}-{sec.name}" if sec.strand and sec.strand != "NONE" else f"{course.code}@{sec.name}"
            data.append(
                {
                    "teacher_id": str(request.user.id),
                    "course_section_id": str(cs.id),
                    "course_id": str(course.id),
                    "course_code": course.code,
                    "course_title": course.title,
                    "cover_image_url": course.cover_image_url,
                    "color_overlay": course.color_overlay,
                    "section_name": sec.name,
                    "strand": sec.strand,
                    "grade_level": sec.grade_level,
                    "course_tag": course_tag,
                    "student_count": cs.student_count,
                    "semester": cs.semester,
                    "school_year": cs.school_year,
                }
            )
        return Response(data)


class CourseSectionDetailView(APIView):
    """Get a single course section by ID."""
    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).select_related("course", "section", "teacher").first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        if request.user.role == User.Role.STUDENT:
            allowed = Enrollment.objects.filter(course_section=course_section, student=request.user, is_active=True).exists()
            if not allowed:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role == User.Role.TEACHER:
            if course_section.teacher_id != request.user.id:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        # Admins can access any course section

        course = course_section.course
        section = course_section.section
        teacher = course_section.teacher

        data = {
            "id": str(course_section.id),
            "course_id": str(course.id),
            "section_id": str(section.id),
            "teacher_id": str(teacher.id) if teacher else None,
            "school_year": course_section.school_year,
            "semester": course_section.semester,
            "is_active": course_section.is_active,
            "course": {
                "id": str(course.id),
                "code": course.code,
                "title": course.title,
                "description": course.description,
                "cover_image_url": course.cover_image_url,
                "color_overlay": course.color_overlay,
                "grade_level": course.grade_level,
                "strand": course.strand,
                "school_year": course.school_year,
                "semester": course.semester,
                "num_weeks": course.num_weeks,
                "is_active": course.is_active,
                "created_at": course.created_at,
                "updated_at": course.updated_at,
            },
            "section": {
                "id": str(section.id),
                "name": section.name,
                "strand": section.strand,
                "grade_level": section.grade_level,
            },
            "teacher": {
                "id": str(teacher.id),
                "first_name": teacher.first_name,
                "last_name": teacher.last_name,
                "full_name": teacher.full_name,
                "email": teacher.email,
                "avatar_url": teacher.avatar_url,
            } if teacher else None,
        }
        return Response(data)


class CourseSectionContentView(APIView):
    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.STUDENT:
            allowed = Enrollment.objects.filter(course_section=course_section, student=request.user, is_active=True).exists()
            if not allowed:
                return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        modules = WeeklyModule.objects.filter(course_section=course_section).order_by("week_number")
        activities = Activity.objects.filter(course_section=course_section).order_by("deadline")
        files = CourseFile.objects.filter(course_section=course_section).order_by("-created_at")
        if request.user.role == User.Role.STUDENT:
            files = files.filter(is_visible=True)
        announcements = Announcement.objects.filter(course_section=course_section).order_by("-created_at")
        quizzes = Quiz.objects.filter(course_section=course_section).order_by("-created_at")

        activities_data = ActivitySerializer(activities, many=True).data
        quizzes_data = QuizSerializer(quizzes, many=True).data

        if request.user.role == User.Role.STUDENT:
            activity_map = {
                str(s.activity_id): s
                for s in Submission.objects.filter(activity__in=activities, student=request.user)
            }
            activity_stats = (
                Submission.objects.filter(activity__in=activities, score__isnull=False)
                .values("activity_id")
                .annotate(lowest=Avg("score"), highest=Avg("score"))
            )
            # We need min/max; SQLite compatibility through Python fallback
            submissions_by_activity = {}
            for sub in Submission.objects.filter(activity__in=activities, score__isnull=False):
                key = str(sub.activity_id)
                submissions_by_activity.setdefault(key, []).append(float(sub.score))
            for item in activities_data:
                sub = activity_map.get(item["id"])
                item["my_submission"] = (
                    SubmissionSerializer(sub).data if sub else None
                )
                scores = submissions_by_activity.get(item["id"], [])
                item["class_stats"] = {
                    "lowest_score": min(scores) if scores else None,
                    "highest_score": max(scores) if scores else None,
                    "average_score": (sum(scores) / len(scores)) if scores else None,
                }

            quiz_attempts = (
                QuizAttempt.objects.filter(quiz__in=quizzes, student=request.user, is_submitted=True)
                .order_by("quiz_id", "-attempt_number")
            )
            in_progress_attempts = (
                QuizAttempt.objects.filter(quiz__in=quizzes, student=request.user, is_submitted=False)
                .order_by("quiz_id", "-attempt_number")
            )
            latest_by_quiz = {}
            for attempt in quiz_attempts:
                key = str(attempt.quiz_id)
                if key not in latest_by_quiz:
                    latest_by_quiz[key] = attempt
            in_progress_by_quiz = {}
            for attempt in in_progress_attempts:
                key = str(attempt.quiz_id)
                if key not in in_progress_by_quiz:
                    in_progress_by_quiz[key] = attempt
            for item in quizzes_data:
                quiz_obj = next((q for q in quizzes if str(q.id) == item["id"]), None)
                attempt = latest_by_quiz.get(item["id"])
                in_progress = in_progress_by_quiz.get(item["id"])
                attempts_used = QuizAttempt.objects.filter(quiz_id=item["id"], student=request.user, is_submitted=True).count()
                attempt_limit = quiz_obj.attempt_limit if quiz_obj else 1
                time_remaining = None
                if in_progress and quiz_obj and quiz_obj.time_limit_minutes:
                    elapsed = (timezone.now() - in_progress.started_at).total_seconds()
                    time_remaining = max(int((quiz_obj.time_limit_minutes * 60) - elapsed), 0)
                if attempt:
                    item["my_attempt"] = {
                        "id": str(attempt.id),
                        "score": float(attempt.score) if attempt.score is not None else None,
                        "max_score": float(attempt.max_score) if attempt.max_score is not None else None,
                        "pending_manual_grading": attempt.pending_manual_grading,
                        "is_submitted": attempt.is_submitted,
                        "attempt_number": attempt.attempt_number,
                        "attempts_used": attempts_used,
                        "attempts_remaining": max(attempt_limit - attempts_used, 0),
                        "attempt_limit": attempt_limit,
                    }
                else:
                    item["my_attempt"] = {
                        "id": None,
                        "score": None,
                        "max_score": None,
                        "pending_manual_grading": False,
                        "is_submitted": False,
                        "attempt_number": 0,
                        "attempts_used": attempts_used,
                        "attempts_remaining": max(attempt_limit - attempts_used, 0),
                        "attempt_limit": attempt_limit,
                    }
                item["my_in_progress_attempt"] = (
                    {
                        "attempt_id": str(in_progress.id),
                        "attempt_number": in_progress.attempt_number,
                        "time_remaining_seconds": time_remaining,
                    }
                    if in_progress
                    else None
                )

        return Response(
            {
                "modules": WeeklyModuleSerializer(modules, many=True).data,
                "activities": activities_data,
                "files": CourseFileSerializer(files, many=True).data,
                "announcements": AnnouncementSerializer(announcements, many=True).data,
                "quizzes": quizzes_data,
            }
        )


class AttendanceOverviewView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _resolve_course_section(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return None, Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.STUDENT:
            allowed = Enrollment.objects.filter(course_section=course_section, student=request.user, is_active=True).exists()
            if not allowed:
                return None, Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role == User.Role.TEACHER:
            if course_section.teacher_id != request.user.id:
                return None, Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role != User.Role.ADMIN:
            return None, Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        return course_section, None

    def _attendance_percentage(self, present_count, absent_count, late_count, excused_count):
        total = present_count + absent_count + late_count + excused_count
        if total <= 0:
            return 0
        score = present_count + excused_count + (late_count * 0.5)
        return int(round((score / total) * 100))

    def get(self, request, pk):
        course_section, denied = self._resolve_course_section(request, pk)
        if denied:
            return denied

        sessions = list(MeetingSession.objects.filter(course_section=course_section).order_by("-date", "-created_at"))
        session_ids = [s.id for s in sessions]
        records_qs = AttendanceRecord.objects.filter(meeting_id__in=session_ids).select_related("student", "meeting")
        records = list(records_qs)

        enrollments = Enrollment.objects.filter(course_section=course_section, is_active=True).select_related("student")
        students = [e.student for e in enrollments if e.student and e.student.role == User.Role.STUDENT]

        if request.user.role == User.Role.STUDENT:
            my_records = [r for r in records if r.student_id == request.user.id]
            present_count = sum(1 for r in my_records if r.status == AttendanceRecord.AttendanceStatus.PRESENT)
            absent_count = sum(1 for r in my_records if r.status == AttendanceRecord.AttendanceStatus.ABSENT)
            late_count = sum(1 for r in my_records if r.status == AttendanceRecord.AttendanceStatus.LATE)
            excused_count = sum(1 for r in my_records if r.status == AttendanceRecord.AttendanceStatus.EXCUSED)
            total_sessions = len(sessions)
            summary = {
                "total_sessions": total_sessions,
                "present_count": present_count,
                "absent_count": absent_count,
                "late_count": late_count,
                "excused_count": excused_count,
                "attendance_percentage": self._attendance_percentage(
                    present_count, absent_count, late_count, excused_count
                ),
            }
            history = []
            by_meeting = {r.meeting_id: r for r in my_records}
            for s in sessions:
                rec = by_meeting.get(s.id)
                history.append(
                    {
                        "meeting_id": str(s.id),
                        "date": s.date,
                        "title": s.title,
                        "status": rec.status if rec else AttendanceRecord.AttendanceStatus.ABSENT,
                        "remarks": rec.remarks if rec else None,
                    }
                )
            return Response(
                {
                    "course_section_id": str(course_section.id),
                    "sessions": MeetingSessionSerializer(sessions, many=True).data,
                    "summary": summary,
                    "history": history,
                    "updated_at": timezone.now(),
                }
            )

        student_rows = []
        for student in students:
            s_records = [r for r in records if r.student_id == student.id]
            present_count = sum(1 for r in s_records if r.status == AttendanceRecord.AttendanceStatus.PRESENT)
            absent_count = sum(1 for r in s_records if r.status == AttendanceRecord.AttendanceStatus.ABSENT)
            late_count = sum(1 for r in s_records if r.status == AttendanceRecord.AttendanceStatus.LATE)
            excused_count = sum(1 for r in s_records if r.status == AttendanceRecord.AttendanceStatus.EXCUSED)

            # Get avatar URL (handles both avatar_url field and avatar FileField)
            avatar_url = None
            if student.avatar_url:
                avatar_url = student.avatar_url
            elif student.avatar:
                avatar_url = request.build_absolute_uri(student.avatar.url)

            student_rows.append(
                {
                    "student_id": str(student.id),
                    "student_name": student.full_name,
                    "student_email": student.email,
                    "avatar_url": avatar_url,
                    "total_sessions": len(sessions),
                    "present_count": present_count,
                    "absent_count": absent_count,
                    "late_count": late_count,
                    "excused_count": excused_count,
                    "attendance_percentage": self._attendance_percentage(
                        present_count, absent_count, late_count, excused_count
                    ),
                }
            )

        return Response(
            {
                "course_section_id": str(course_section.id),
                "sessions": MeetingSessionSerializer(sessions, many=True).data,
                "students": student_rows,
                "records": AttendanceRecordSerializer(records, many=True).data,
                "updated_at": timezone.now(),
            }
        )


class AttendanceSessionCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        date = request.data.get("date")
        title = request.data.get("title")
        if not date or not title:
            return Response({"detail": "date and title are required."}, status=status.HTTP_400_BAD_REQUEST)

        session = MeetingSession.objects.create(
            course_section=course_section,
            date=date,
            title=title,
            created_by=request.user,
        )
        enrolled_students = User.objects.filter(
            enrollments__course_section=course_section,
            enrollments__is_active=True,
            role=User.Role.STUDENT,
        ).distinct()
        AttendanceRecord.objects.bulk_create(
            [
                AttendanceRecord(
                    meeting=session,
                    student=student,
                    status=AttendanceRecord.AttendanceStatus.ABSENT,
                    marked_by=request.user,
                )
                for student in enrolled_students
            ]
        )
        _recompute_course_section_grades(course_section)
        return Response(MeetingSessionSerializer(session).data, status=status.HTTP_201_CREATED)


class AttendanceSessionDeleteView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def delete(self, request, pk):
        session = MeetingSession.objects.select_related("course_section").filter(id=pk).first()
        if not session:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and session.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        course_section = session.course_section
        session.delete()
        _recompute_course_section_grades(course_section)
        return Response(status=status.HTTP_204_NO_CONTENT)


class AttendanceRecordBulkUpdateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        session = MeetingSession.objects.select_related("course_section").filter(id=pk).first()
        if not session:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and session.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        bulk_action = request.data.get("bulk_action")
        if bulk_action == "mark_all_present":
            AttendanceRecord.objects.filter(meeting=session).update(
                status=AttendanceRecord.AttendanceStatus.PRESENT,
                marked_by=request.user,
                updated_at=timezone.now(),
            )
            _recompute_course_section_grades(session.course_section)
            return Response({"ok": True})
        if bulk_action == "clear_all":
            AttendanceRecord.objects.filter(meeting=session).update(
                status=AttendanceRecord.AttendanceStatus.ABSENT,
                remarks=None,
                marked_by=request.user,
                updated_at=timezone.now(),
            )
            _recompute_course_section_grades(session.course_section)
            return Response({"ok": True})

        records = request.data.get("records") or []
        if not isinstance(records, list):
            return Response({"detail": "records must be a list."}, status=status.HTTP_400_BAD_REQUEST)

        valid_statuses = set(AttendanceRecord.AttendanceStatus.values)
        for row in records:
            student_id = row.get("student_id")
            status_value = row.get("status")
            if not student_id or status_value not in valid_statuses:
                continue
            AttendanceRecord.objects.update_or_create(
                meeting=session,
                student_id=student_id,
                defaults={
                    "status": status_value,
                    "remarks": row.get("remarks"),
                    "marked_by": request.user,
                },
            )
        _recompute_course_section_grades(session.course_section)
        return Response({"ok": True})


class CourseSectionGradesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        enrollments = Enrollment.objects.filter(course_section=course_section, is_active=True).select_related("student")
        rows = []
        for e in enrollments:
            _recompute_enrollment_grade(e)
            grade = float(e.final_grade) if e.final_grade is not None else None
            rows.append(
                {
                    "enrollment_id": str(e.id),
                    "student_id": str(e.student_id),
                    "student_name": e.student.full_name,
                    "student_email": e.student.email,
                    "final_grade": grade,
                    "final_grade_letter": _letter_grade(Decimal(str(grade))) if grade is not None else None,
                    "grade_overridden": e.manual_final_grade is not None,
                    "manual_final_grade": float(e.manual_final_grade) if e.manual_final_grade is not None else None,
                }
            )
        return Response(rows)


class CourseSectionGradebookView(APIView):
    """Returns comprehensive gradebook data including activities and quizzes for each student."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        course_section = CourseSection.objects.filter(id=pk).first()
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Get all published activities and quizzes
        activities = list(
            Activity.objects.filter(course_section=course_section, is_published=True)
            .order_by("created_at", "title")
            .values("id", "title", "points", "deadline", "created_at")
        )
        quizzes = list(
            Quiz.objects.filter(course_section=course_section, is_published=True)
            .order_by("created_at", "title")
            .values("id", "title", "score_selection_policy", "close_at", "created_at")
        )

        # Get quiz max scores
        quiz_points_raw = (
            QuizQuestion.objects.filter(quiz_id__in=[q["id"] for q in quizzes])
            .values("quiz_id")
            .annotate(total=Sum("points"))
        )
        quiz_points = {str(row["quiz_id"]): Decimal(str(row["total"] or 0)) for row in quiz_points_raw}

        # Get all enrollments (both active and inactive)
        all_enrollments = list(
            Enrollment.objects.filter(course_section=course_section)
            .select_related("student")
            .order_by("-is_active", "student__last_name", "student__first_name", "student__email")
        )
        student_ids = [e.student_id for e in all_enrollments]

        # Get all submissions for these activities
        activity_ids = [a["id"] for a in activities]
        submissions = Submission.objects.filter(
            activity_id__in=activity_ids,
            student_id__in=student_ids
        ).values("id", "activity_id", "student_id", "score", "status", "submitted_at", "graded_at")

        # Build submission maps: {(student_id, activity_id): submission_data}
        submission_map: dict[tuple, dict] = {}
        for s in submissions:
            key = (str(s["student_id"]), str(s["activity_id"]))
            submission_map[key] = s

        # Get all quiz attempts
        quiz_ids = [q["id"] for q in quizzes]
        quiz_attempts = QuizAttempt.objects.filter(
            quiz_id__in=quiz_ids,
            student_id__in=student_ids,
            is_submitted=True
        ).values("id", "quiz_id", "student_id", "score", "max_score", "attempt_number", "submitted_at", "pending_manual_grading")

        # Build quiz attempt maps: {(student_id, quiz_id): [attempts]}
        attempts_by_quiz: dict[tuple, list] = {}
        for a in quiz_attempts:
            key = (str(a["student_id"]), str(a["quiz_id"]))
            if key not in attempts_by_quiz:
                attempts_by_quiz[key] = []
            attempts_by_quiz[key].append(a)

        # Build student data
        active_students = []
        inactive_students = []

        for enrollment in all_enrollments:
            _recompute_enrollment_grade(enrollment)
            student_id = str(enrollment.student_id)
            enrolled_at = enrollment.enrolled_at

            # Build activity grades for this student
            activity_grades = []
            for activity in activities:
                activity_id = str(activity["id"])
                activity_deadline = activity["deadline"]

                # Check if activity deadline is before enrollment (pre-enrollment exclusion)
                if activity_deadline and enrolled_at:
                    from datetime import datetime
                    if isinstance(activity_deadline, str):
                        activity_deadline = datetime.fromisoformat(activity_deadline.replace('Z', '+00:00'))
                    if isinstance(enrolled_at, str):
                        enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
                    else:
                        enrolled_at_dt = enrolled_at
                    if activity_deadline < enrolled_at_dt:
                        activity_grades.append({
                            "activity_id": activity_id,
                            "title": activity["title"],
                            "points": float(activity["points"]),
                            "deadline": activity["deadline"],
                            "score": None,
                            "status": None,
                            "is_late": False,
                            "is_excused": False,
                            "graded_at": None,
                            "is_na": True,  # Pre-enrollment exclusion
                        })
                        continue

                sub = submission_map.get((student_id, activity_id))
                if sub:
                    is_late = False
                    if sub["status"] == Submission.SubmissionStatus.LATE:
                        is_late = True
                    score = float(sub["score"]) if sub["score"] is not None else None
                    graded_at = sub["graded_at"].isoformat() if sub["graded_at"] else None

                    # Determine status
                    status = sub["status"]
                    if score is None and sub["status"] == Submission.SubmissionStatus.SUBMITTED:
                        status = "submitted"  # Needs grading

                    activity_grades.append({
                        "activity_id": activity_id,
                        "title": activity["title"],
                        "points": float(activity["points"]),
                        "deadline": activity["deadline"],
                        "score": score,
                        "status": status if status else "not_submitted",
                        "is_late": is_late,
                        "is_excused": False,
                        "graded_at": graded_at,
                        "is_na": False,
                    })
                else:
                    activity_grades.append({
                        "activity_id": activity_id,
                        "title": activity["title"],
                        "points": float(activity["points"]),
                        "deadline": activity["deadline"],
                        "score": None,
                        "status": "not_submitted",
                        "is_late": False,
                        "is_excused": False,
                        "graded_at": None,
                        "is_na": False,
                    })

            # Build quiz grades for this student
            quiz_grades = []
            for quiz in quizzes:
                quiz_id = str(quiz["id"])
                max_score = float(quiz_points.get(quiz_id, 0))
                close_at = quiz["close_at"]

                # Check if quiz close_at is before enrollment (pre-enrollment exclusion)
                if close_at and enrolled_at:
                    from datetime import datetime
                    if isinstance(close_at, str):
                        close_at_dt = datetime.fromisoformat(close_at.replace('Z', '+00:00'))
                    else:
                        close_at_dt = close_at
                    if isinstance(enrolled_at, str):
                        enrolled_at_dt = datetime.fromisoformat(enrolled_at.replace('Z', '+00:00'))
                    else:
                        enrolled_at_dt = enrolled_at
                    if close_at_dt < enrolled_at_dt:
                        quiz_grades.append({
                            "quiz_id": quiz_id,
                            "title": quiz["title"],
                            "max_score": max_score,
                            "close_at": quiz["close_at"],
                            "score": None,
                            "attempts": 0,
                            "max_attempts": 0,
                            "is_late": False,
                            "is_na": True,
                        })
                        continue

                attempts = attempts_by_quiz.get((student_id, quiz_id), [])
                if attempts:
                    # Calculate score based on policy
                    scores = [float(a["score"]) for a in attempts if a["score"] is not None]
                    if scores:
                        policy = quiz.get("score_selection_policy", "highest")
                        if policy == "highest":
                            score = max(scores)
                        else:  # latest
                            # Sort by attempt_number descending and get the last submitted
                            sorted_attempts = sorted(attempts, key=lambda x: x["attempt_number"], reverse=True)
                            score = float(sorted_attempts[0]["score"]) if sorted_attempts[0]["score"] is not None else None
                    else:
                        score = None

                    # Check if any attempt is pending grading
                    pending_grading = any(a.get("pending_manual_grading", False) for a in attempts)

                    quiz_grades.append({
                        "quiz_id": quiz_id,
                        "title": quiz["title"],
                        "max_score": max_score,
                        "close_at": quiz["close_at"],
                        "score": score,
                        "attempts": len(attempts),
                        "max_attempts": 0,  # TODO: add max_attempts to Quiz model if needed
                        "is_late": False,  # TODO: check if submitted after close_at
                        "is_na": False,
                        "pending_grading": pending_grading,
                    })
                else:
                    quiz_grades.append({
                        "quiz_id": quiz_id,
                        "title": quiz["title"],
                        "max_score": max_score,
                        "close_at": quiz["close_at"],
                        "score": None,
                        "attempts": 0,
                        "max_attempts": 0,
                        "is_late": False,
                        "is_na": False,
                    })

            student_data = {
                "enrollment_id": str(enrollment.id),
                "student_id": student_id,
                "student_name": enrollment.student.full_name,
                "student_email": enrollment.student.email,
                "student_avatar": enrollment.student.avatar_url if hasattr(enrollment.student, 'avatar_url') else None,
                "enrolled_at": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None,
                "is_active": enrollment.is_active,
                "grades": {
                    "activities": activity_grades,
                    "quizzes": quiz_grades,
                },
                "final_grade": float(enrollment.final_grade) if enrollment.final_grade is not None else None,
                "final_grade_letter": _letter_grade(enrollment.final_grade),
                "grade_overridden": enrollment.manual_final_grade is not None,
                "manual_final_grade": float(enrollment.manual_final_grade) if enrollment.manual_final_grade is not None else None,
            }

            if enrollment.is_active:
                active_students.append(student_data)
            else:
                inactive_students.append(student_data)

        # Build items list (column headers)
        items = {
            "activities": [
                {
                    "id": str(a["id"]),
                    "title": a["title"],
                    "type": "activity",
                    "max_points": float(a["points"]),
                    "deadline": a["deadline"].isoformat() if a["deadline"] else None,
                    "created_at": a["created_at"].isoformat() if a["created_at"] else None,
                }
                for a in activities
            ],
            "quizzes": [
                {
                    "id": str(q["id"]),
                    "title": q["title"],
                    "type": "quiz",
                    "max_points": float(quiz_points.get(str(q["id"]), 0)),
                    "close_at": q["close_at"].isoformat() if q["close_at"] else None,
                    "created_at": q["created_at"].isoformat() if q["created_at"] else None,
                }
                for q in quizzes
            ],
        }

        # Build summary statistics for each item
        activity_summary = []
        for activity in activities:
            activity_id = str(activity["id"])
            scores = []
            missing_count = 0
            needs_grading_count = 0

            for student in active_students:
                grade = next((g for g in student["grades"]["activities"] if g["activity_id"] == activity_id), None)
                if grade:
                    if grade.get("is_na"):
                        continue  # Exclude from stats
                    if grade["score"] is not None:
                        scores.append(grade["score"])
                    elif grade["status"] == "not_submitted":
                        missing_count += 1
                    elif grade["status"] == "submitted":
                        needs_grading_count += 1
                    elif grade["status"] == "late":
                        needs_grading_count += 1  # Late submissions also need grading if no score

            activity_summary.append({
                "activity_id": activity_id,
                "avg_score": float(sum(scores) / len(scores)) if scores else None,
                "high_score": float(max(scores)) if scores else None,
                "low_score": float(min(scores)) if scores else None,
                "missing_count": missing_count,
                "needs_grading_count": needs_grading_count,
            })

        quiz_summary = []
        for quiz in quizzes:
            quiz_id = str(quiz["id"])
            scores = []
            missing_count = 0
            needs_grading_count = 0

            for student in active_students:
                grade = next((g for g in student["grades"]["quizzes"] if g["quiz_id"] == quiz_id), None)
                if grade:
                    if grade.get("is_na"):
                        continue
                    if grade["score"] is not None:
                        scores.append(grade["score"])
                    elif grade["attempts"] == 0:
                        missing_count += 1
                    if grade.get("pending_grading"):
                        needs_grading_count += 1

            quiz_summary.append({
                "quiz_id": quiz_id,
                "avg_score": float(sum(scores) / len(scores)) if scores else None,
                "high_score": float(max(scores)) if scores else None,
                "low_score": float(min(scores)) if scores else None,
                "missing_count": missing_count,
                "needs_grading_count": needs_grading_count,
            })

        return Response({
            "students": active_students,
            "inactive_students": inactive_students,
            "items": items,
            "summary": {
                "activities": activity_summary,
                "quizzes": quiz_summary,
            },
        })


class CourseSectionGradesExportCSVView(APIView):
    """Export grades as CSV or XLSX with optional filtering."""
    permission_classes = [permissions.IsAuthenticated]

    def _format_num(self, value: Decimal | float | int | None) -> str:
        if value is None:
            return ""
        dec = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{dec:.2f}"

    def _stream_text(self, text: str, chunk_size: int = 8192):
        for idx in range(0, len(text), chunk_size):
            yield text[idx : idx + chunk_size]

    def _resolve_course_section(self, *, pk=None, course_id=None, section_id=None):
        target_id = section_id or pk
        if not target_id:
            return None
        course_section = CourseSection.objects.select_related("course", "section").filter(id=target_id).first()
        if not course_section:
            return None
        if course_id and str(course_section.course_id) != str(course_id):
            return None
        return course_section

    def _generate_csv(self, course_section, enrollments, activities, quizzes, submission_map, quiz_attempt_map, include_inactive=False, include_student_id=False, include_enrolled_at=False):
        """Generate CSV export."""
        headers = ["Student Name", "Student Email"]
        if include_student_id:
            headers.append("Student ID")
        if include_enrolled_at:
            headers.append("Enrolled At")
        headers.extend(["Section", "Total Grade", "Grade Letter"])
        headers.extend([f"Activity: {a.title}" for a in activities])
        headers.extend([f"Quiz: {q.title}" for q in quizzes])

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)

        for enrollment in enrollments:
            if not enrollment.is_active and not include_inactive:
                continue
            computed_total = _compute_enrollment_grade(enrollment)
            row = [
                enrollment.student.full_name,
                enrollment.student.email,
            ]
            if include_student_id:
                row.append(str(enrollment.student_id))
            if include_enrolled_at:
                row.append(enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else "")
            row.extend([
                course_section.section.name,
                self._format_num(computed_total),
                _letter_grade(computed_total) if computed_total else "",
            ])
            for activity in activities:
                sub = submission_map.get((str(enrollment.student_id), str(activity.id)))
                if sub and sub.score is not None:
                    row.append(f"{self._format_num(sub.score)}/{self._format_num(activity.points)}")
                else:
                    row.append("")
            for quiz in quizzes:
                attempt = quiz_attempt_map.get((str(enrollment.student_id), str(quiz.id)))
                if attempt and attempt.score is not None and attempt.max_score is not None:
                    row.append(f"{self._format_num(attempt.score)}/{self._format_num(attempt.max_score)}")
                elif attempt and attempt.score is not None:
                    row.append(self._format_num(attempt.score))
                else:
                    row.append("")
            writer.writerow(row)

        return "\ufeff" + output.getvalue()

    def _generate_xlsx(self, course_section, enrollments, activities, quizzes, submission_map, quiz_attempt_map, include_inactive=False, include_student_id=False, include_enrolled_at=False):
        """Generate XLSX export with conditional formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"

        # Define styles
        header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        missing_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red
        needs_grading_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        passing_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Build headers
        headers = ["Student Name", "Student Email"]
        if include_student_id:
            headers.append("Student ID")
        if include_enrolled_at:
            headers.append("Enrolled At")
        headers.extend(["Section", "Total Grade", "Grade Letter"])
        headers.extend([f"Activity: {a.title}" for a in activities])
        headers.extend([f"Quiz: {q.title}" for q in quizzes])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 25  # Student Name
        ws.column_dimensions['B'].width = 30  # Student Email
        if include_student_id:
            ws.column_dimensions['C'].width = 36  # Student ID
        if include_enrolled_at:
            extra_col = 1 if include_student_id else 0
            ws.column_dimensions[get_column_letter(3 + extra_col)].width = 20  # Enrolled At

        # Data rows
        row_num = 2
        for enrollment in enrollments:
            if not enrollment.is_active and not include_inactive:
                continue
            computed_total = _compute_enrollment_grade(enrollment)

            col = 1
            ws.cell(row=row_num, column=col, value=enrollment.student.full_name).border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=enrollment.student.email).border = thin_border
            col += 1
            if include_student_id:
                ws.cell(row=row_num, column=col, value=str(enrollment.student_id)).border = thin_border
                col += 1
            if include_enrolled_at:
                ws.cell(row=row_num, column=col, value=enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else "").border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value=course_section.section.name).border = thin_border
            col += 1
            grade_cell = ws.cell(row=row_num, column=col, value=float(computed_total) if computed_total else "")
            grade_cell.border = thin_border
            if computed_total and computed_total >= Decimal("70"):
                grade_cell.fill = passing_fill
            col += 1
            ws.cell(row=row_num, column=col, value=_letter_grade(computed_total) if computed_total else "").border = thin_border
            col += 1

            # Activity scores
            for activity in activities:
                sub = submission_map.get((str(enrollment.student_id), str(activity.id)))
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                if sub and sub.score is not None:
                    cell.value = f"{float(sub.score):.1f}/{float(activity.points):.0f}"
                    if sub.score / activity.points >= Decimal("0.7"):
                        cell.fill = passing_fill
                else:
                    cell.value = ""
                    cell.fill = missing_fill
                col += 1

            # Quiz scores
            for quiz in quizzes:
                attempt = quiz_attempt_map.get((str(enrollment.student_id), str(quiz.id)))
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                if attempt and attempt.score is not None and attempt.max_score is not None:
                    cell.value = f"{float(attempt.score):.1f}/{float(attempt.max_score):.0f}"
                    if attempt.score / attempt.max_score >= Decimal("0.7"):
                        cell.fill = passing_fill
                elif attempt and attempt.score is not None:
                    cell.value = f"{float(attempt.score):.1f}"
                else:
                    cell.value = ""
                    cell.fill = missing_fill
                col += 1

            row_num += 1

        # Freeze first row (header)
        ws.freeze_panes = 'A2'

        # Write to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def get(self, request, pk=None, course_id=None, section_id=None):
        course_section = self._resolve_course_section(pk=pk, course_id=course_id, section_id=section_id)
        if not course_section:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Query parameters
        export_format = request.query_params.get("format", "csv").lower()
        scope = request.query_params.get("scope", "all").lower()  # all, activities, quizzes, final_only
        include_inactive = request.query_params.get("include_inactive", "false").lower() == "true"
        include_student_id = request.query_params.get("include_student_id", "false").lower() == "true"
        include_enrolled_at = request.query_params.get("include_enrolled_at", "false").lower() == "true"

        activities = list(
            Activity.objects.filter(course_section=course_section, is_published=True).order_by("created_at", "title")
        )
        quizzes = list(
            Quiz.objects.filter(course_section=course_section, is_published=True).order_by("created_at", "title")
        )

        # Filter by scope
        if scope == "activities":
            quizzes = []
        elif scope == "quizzes":
            activities = []
        elif scope == "final_only":
            activities = []
            quizzes = []

        enrollments = list(
            Enrollment.objects.filter(course_section=course_section)
            .select_related("student")
            .order_by("-is_active", "student__last_name", "student__first_name", "student__email")
        )
        student_ids = [e.student_id for e in enrollments]

        submission_map: dict[tuple[str, str], Submission] = {}
        if activities and student_ids:
            submissions = Submission.objects.filter(
                activity_id__in=[a.id for a in activities],
                student_id__in=student_ids,
            )
            submission_map = {(str(s.student_id), str(s.activity_id)): s for s in submissions}

        quiz_attempt_map: dict[tuple[str, str], QuizAttempt] = {}
        if quizzes and student_ids:
            quiz_attempts = (
                QuizAttempt.objects.filter(
                    quiz_id__in=[q.id for q in quizzes],
                    student_id__in=student_ids,
                    is_submitted=True,
                    score__isnull=False,
                )
                .order_by("student_id", "quiz_id", "-attempt_number", "-submitted_at")
            )
            for attempt in quiz_attempts:
                key = (str(attempt.student_id), str(attempt.quiz_id))
                if key not in quiz_attempt_map:
                    quiz_attempt_map[key] = attempt

        base_name = f"{course_section.course.code}_{course_section.section.name}_grades".replace(" ", "_")

        if export_format == "xlsx":
            xlsx_data = self._generate_xlsx(
                course_section, enrollments, activities, quizzes,
                submission_map, quiz_attempt_map,
                include_inactive, include_student_id, include_enrolled_at
            )
            if xlsx_data is None:
                return Response({"detail": "XLSX generation not available."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            response = HttpResponse(
                xlsx_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{base_name}.xlsx"'
            return response

        # Default to CSV
        csv_text = self._generate_csv(
            course_section, enrollments, activities, quizzes,
            submission_map, quiz_attempt_map,
            include_inactive, include_student_id, include_enrolled_at
        )

        logger.info(
            "grade_csv_export user_id=%s section_id=%s students=%s activities=%s quizzes=%s",
            request.user.id,
            course_section.id,
            len([e for e in enrollments if e.is_active]),
            len(activities),
            len(quizzes),
        )

        response = StreamingHttpResponse(self._stream_text(csv_text), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
        return response


class EnrollmentGradeOverrideView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        enrollment = Enrollment.objects.select_related("course_section").filter(id=pk).first()
        if not enrollment:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and enrollment.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        value = request.data.get("manual_final_grade", None)
        if value in ["", None]:
            enrollment.manual_final_grade = None
        else:
            try:
                enrollment.manual_final_grade = Decimal(str(value))
            except Exception:
                return Response({"detail": "manual_final_grade must be numeric."}, status=status.HTTP_400_BAD_REQUEST)

        _recompute_enrollment_grade(enrollment)
        return Response(
            {
                "enrollment_id": str(enrollment.id),
                "manual_final_grade": float(enrollment.manual_final_grade) if enrollment.manual_final_grade is not None else None,
                "final_grade": float(enrollment.final_grade) if enrollment.final_grade is not None else None,
                "final_grade_letter": _letter_grade(enrollment.final_grade),
                "grade_overridden": enrollment.manual_final_grade is not None,
            }
        )


class ActivitySubmitView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can submit activities."}, status=status.HTTP_403_FORBIDDEN)

        activity = Activity.objects.filter(id=pk).first()
        if not activity:
            return Response({"detail": "Activity not found."}, status=status.HTTP_404_NOT_FOUND)

        enrolled = Enrollment.objects.filter(
            course_section=activity.course_section,
            student=request.user,
            is_active=True,
        ).exists()
        if not enrolled:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Check attempt limit
        attempt_limit = activity.attempt_limit or 1
        existing_submissions = Submission.objects.filter(activity=activity, student=request.user).count()
        if existing_submissions >= attempt_limit:
            return Response({"detail": f"Attempt limit reached. You have used {existing_submissions} of {attempt_limit} attempts."}, status=status.HTTP_400_BAD_REQUEST)

        allowed = activity.allowed_file_types or ["all"]
        allowed_set = set([a.lower() for a in allowed if isinstance(a, str)])
        allow_all = "all" in allowed_set
        allow_text = allow_all or "text" in allowed_set
        allow_image = allow_all or "image" in allowed_set
        allow_pdf = allow_all or "pdf" in allowed_set

        text_content = request.data.get("text_content")
        if text_content and not allow_text:
            return Response({"detail": "Text submission is not allowed for this activity."}, status=status.HTTP_400_BAD_REQUEST)

        existing_urls = request.data.get("file_urls") or []
        if not isinstance(existing_urls, list):
            existing_urls = []

        uploaded_urls = []
        files = request.FILES.getlist("files")
        for file_obj in files:
            ctype = (file_obj.content_type or "").lower()
            is_image = ctype.startswith("image/")
            is_pdf = ctype == "application/pdf" or file_obj.name.lower().endswith(".pdf")
            if (is_image and not allow_image) or (is_pdf and not allow_pdf) or (not is_image and not is_pdf and not allow_all):
                return Response({"detail": f"File type not allowed: {file_obj.name}"}, status=status.HTTP_400_BAD_REQUEST)
            path = default_storage.save(f"submissions/{request.user.id}/{timezone.now().timestamp()}_{file_obj.name}", file_obj)
            uploaded_urls.append(default_storage.url(path))

        now = timezone.now()
        status_value = Submission.SubmissionStatus.SUBMITTED
        if activity.deadline and now > activity.deadline:
            status_value = Submission.SubmissionStatus.LATE

        # Create new submission with incremented attempt_number
        next_attempt_number = existing_submissions + 1
        submission = Submission.objects.create(
            activity=activity,
            student=request.user,
            attempt_number=next_attempt_number,
            file_urls=existing_urls + uploaded_urls,
            text_content=text_content,
            submitted_at=now,
            status=status_value,
        )
        enrollment = Enrollment.objects.filter(course_section=activity.course_section, student=request.user, is_active=True).first()
        if enrollment:
            _recompute_enrollment_grade(enrollment)
        _sync_student_activity_items(request.user)
        return Response(SubmissionSerializer(submission).data)


class ActivityMySubmissionView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can view own submission here."}, status=status.HTTP_403_FORBIDDEN)
        # Return the latest submission
        submission = Submission.objects.filter(activity_id=pk, student=request.user).order_by("-attempt_number").first()
        # Also return all submissions for the student's reference
        all_submissions = Submission.objects.filter(activity_id=pk, student=request.user).order_by("-attempt_number")
        activity = Activity.objects.filter(id=pk).first()
        attempt_limit = activity.attempt_limit if activity else 1
        attempts_used = all_submissions.count()
        return Response({
            "submission": SubmissionSerializer(submission).data if submission else None,
            "all_submissions": SubmissionSerializer(all_submissions, many=True).data,
            "attempt_limit": attempt_limit,
            "attempts_used": attempts_used,
            "attempts_remaining": max(attempt_limit - attempts_used, 0),
        })


class ActivitySubmissionsForTeacherView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        activity = Activity.objects.select_related("course_section").filter(id=pk).first()
        if not activity:
            return Response({"detail": "Activity not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and activity.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Get all submissions ordered by attempt number
        all_submissions = Submission.objects.filter(activity=activity).select_related("student").order_by("student_id", "-attempt_number")

        # Group submissions by student
        submissions_by_student: dict[str, list[Submission]] = {}
        for s in all_submissions:
            student_id = str(s.student_id)
            if student_id not in submissions_by_student:
                submissions_by_student[student_id] = []
            submissions_by_student[student_id].append(s)

        # Get latest submission per student (first in the list due to ordering)
        latest_by_student = {student_id: submissions[0] for student_id, submissions in submissions_by_student.items()}

        enrolled_students = User.objects.filter(
            enrollments__course_section=activity.course_section,
            enrollments__is_active=True,
            role=User.Role.STUDENT,
        ).distinct().order_by("last_name", "first_name")

        payload = []
        for student in enrolled_students:
            student_id = str(student.id)
            all_student_subs = submissions_by_student.get(student_id, [])
            latest_sub = latest_by_student.get(student_id)

            # Calculate best score based on policy
            best_score = None
            if all_student_subs:
                scores = [s.score for s in all_student_subs if s.score is not None]
                if scores:
                    if activity.score_selection_policy == Activity.ScorePolicy.HIGHEST:
                        best_score = max(scores)
                    else:
                        best_score = scores[-1]  # Latest score

            # Base submission data
            row = SubmissionSerializer(latest_sub).data if latest_sub else {
                "id": None,
                "activity_id": str(activity.id),
                "student_id": str(student.id),
                "attempt_number": None,
                "file_urls": [],
                "text_content": None,
                "status": Submission.SubmissionStatus.NOT_SUBMITTED,
                "score": None,
                "feedback": None,
                "submitted_at": None,
                "graded_at": None,
                "created_at": None,
                "updated_at": None,
            }
            row["enrollment_status"] = "enrolled"
            row["student_name"] = student.full_name
            row["student_email"] = student.email
            row["attempt_limit"] = activity.attempt_limit
            row["attempts_used"] = len(all_student_subs)
            row["attempts_remaining"] = max(activity.attempt_limit - len(all_student_subs), 0)
            row["all_submissions"] = SubmissionSerializer(all_student_subs, many=True).data
            row["best_score"] = best_score
            row["score_selection_policy"] = activity.score_selection_policy
            payload.append(row)
        return Response(payload)


class ActivitySubmissionGradeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        submission = Submission.objects.select_related("activity__course_section").filter(id=pk).first()
        if not submission:
            return Response({"detail": "Submission not found."}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role == User.Role.TEACHER and submission.activity.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Check if this is a new grade being set
        was_ungraded = submission.score is None

        serializer = SubmissionGradeSerializer(submission, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        graded = serializer.save(graded_at=timezone.now())
        if graded.score is not None:
            graded.status = Submission.SubmissionStatus.GRADED
            graded.save(update_fields=["status", "updated_at"])
            enrollment = Enrollment.objects.filter(
                course_section=graded.activity.course_section,
                student=graded.student,
                is_active=True,
            ).first()
            if enrollment:
                _recompute_enrollment_grade(enrollment)

            # Send push notification to student when grade is released
            if was_ungraded:
                self._send_grade_notification(graded)

        return Response(SubmissionSerializer(graded).data)

    def _send_grade_notification(self, submission: Submission):
        """Send push notification to student when grade is released."""
        from .push_notifications import send_push_notification_to_users

        try:
            activity = submission.activity
            student = submission.student

            # Create in-app notification
            Notification.objects.create(
                recipient=student,
                type=Notification.NotificationType.GRADE_RELEASED,
                title=f"Grade Released: {activity.title}",
                body=f"Your submission for '{activity.title}' has been graded. Score: {submission.score}/{activity.points}",
                course_section=activity.course_section,
                activity=activity,
            )

            # Send push notification
            data = {
                "type": "grade_released",
                "activity_id": str(activity.id),
                "course_section_id": str(activity.course_section_id),
            }

            send_push_notification_to_users(
                user_ids=[str(student.id)],
                title=f"Grade Released: {activity.title}",
                body=f"Your submission has been graded. Score: {submission.score}/{activity.points}",
                data=data,
            )
        except Exception as e:
            logger.warning(f"Failed to send grade notification: {e}")


class QuizTakeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _get_time_remaining(self, attempt: QuizAttempt):
        quiz = attempt.quiz
        if not quiz.time_limit_minutes:
            return None
        elapsed = (timezone.now() - attempt.started_at).total_seconds()
        remaining = int((quiz.time_limit_minutes * 60) - elapsed)
        return max(remaining, 0)

    def _auto_finalize_attempt(self, attempt: QuizAttempt):
        if attempt.is_submitted:
            return attempt

        questions = {str(q.id): q for q in QuizQuestion.objects.filter(quiz=attempt.quiz)}
        answers_by_qid = {str(a.question_id): a for a in QuizAnswer.objects.filter(attempt=attempt)}
        max_score = sum(float(q.points) for q in questions.values())
        total_score = 0.0
        pending_manual = False

        for qid, question in questions.items():
            ans = answers_by_qid.get(qid)
            if not ans:
                if question.question_type == QuizQuestion.QuestionType.ESSAY:
                    pending_manual = True
                    QuizAnswer.objects.create(
                        attempt=attempt,
                        question=question,
                        needs_manual_grading=True,
                    )
                continue

            if question.question_type in [QuizQuestion.QuestionType.MULTIPLE_CHOICE, QuizQuestion.QuestionType.TRUE_FALSE]:
                is_correct = bool(ans.selected_choice and ans.selected_choice.is_correct)
                points_awarded = float(question.points) if is_correct else 0.0
                ans.is_correct = is_correct
                ans.points_awarded = points_awarded
                ans.needs_manual_grading = False
                ans.graded_at = timezone.now()
                ans.save(update_fields=["is_correct", "points_awarded", "needs_manual_grading", "graded_at"])
                total_score += points_awarded
            else:
                pending_manual = True
                ans.needs_manual_grading = True
                ans.is_correct = None
                ans.points_awarded = None
                ans.save(update_fields=["needs_manual_grading", "is_correct", "points_awarded"])

        attempt.is_submitted = True
        attempt.submitted_at = timezone.now()
        attempt.max_score = max_score
        attempt.score = total_score
        attempt.pending_manual_grading = pending_manual
        attempt.save(update_fields=["is_submitted", "submitted_at", "max_score", "score", "pending_manual_grading"])
        enrollment = Enrollment.objects.filter(
            course_section=attempt.quiz.course_section,
            student=attempt.student,
            is_active=True,
        ).first()
        if enrollment:
            _recompute_enrollment_grade(enrollment)
        _sync_student_activity_items(attempt.student)
        return attempt

    def get(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can take quizzes."}, status=status.HTTP_403_FORBIDDEN)

        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)

        enrolled = Enrollment.objects.filter(course_section=quiz.course_section, student=request.user, is_active=True).exists()
        if not enrolled:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        submitted_count = QuizAttempt.objects.filter(quiz=quiz, student=request.user, is_submitted=True).count()
        if submitted_count >= quiz.attempt_limit:
            return Response({"detail": "Attempt limit reached."}, status=status.HTTP_400_BAD_REQUEST)

        open_attempt = (
            QuizAttempt.objects.filter(quiz=quiz, student=request.user, is_submitted=False)
            .order_by("-attempt_number")
            .first()
        )
        if open_attempt:
            attempt = open_attempt
        else:
            attempt = QuizAttempt.objects.create(
                quiz=quiz,
                student=request.user,
                attempt_number=submitted_count + 1,
                is_submitted=False,
            )

        remaining = self._get_time_remaining(attempt)
        if remaining is not None and remaining <= 0:
            self._auto_finalize_attempt(attempt)
            return Response({"detail": "Quiz time has ended and your attempt was auto-submitted."}, status=status.HTTP_400_BAD_REQUEST)

        questions = QuizQuestion.objects.filter(quiz=quiz).prefetch_related("choices").order_by("sort_order")
        existing_answers = QuizAnswer.objects.filter(attempt=attempt)
        answers_payload = []
        for a in existing_answers:
            answers_payload.append(
                {
                    "question_id": str(a.question_id),
                    "selected_choice_id": str(a.selected_choice_id) if a.selected_choice_id else None,
                    "text_answer": a.text_answer,
                }
            )
        return Response(
            {
                "quiz": QuizSerializer(quiz).data,
                "questions": QuizQuestionStudentSerializer(questions, many=True).data,
                "attempt_id": str(attempt.id),
                "attempt_number": attempt.attempt_number,
                "time_remaining_seconds": remaining,
                "answers": answers_payload,
            }
        )


class QuizSubmitAttemptView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can submit quizzes."}, status=status.HTTP_403_FORBIDDEN)

        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)

        enrolled = Enrollment.objects.filter(course_section=quiz.course_section, student=request.user, is_active=True).exists()
        if not enrolled:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        submitted_attempts = QuizAttempt.objects.filter(quiz=quiz, student=request.user, is_submitted=True).count()
        if submitted_attempts >= quiz.attempt_limit:
            return Response({"detail": "Attempt limit reached."}, status=status.HTTP_400_BAD_REQUEST)

        attempt_id = request.data.get("attempt_id")
        attempt = None
        if attempt_id:
            attempt = QuizAttempt.objects.filter(
                id=attempt_id,
                quiz=quiz,
                student=request.user,
                is_submitted=False,
            ).first()

        if not attempt:
            attempt = (
                QuizAttempt.objects.filter(quiz=quiz, student=request.user, is_submitted=False)
                .order_by("-attempt_number")
                .first()
            )
        if not attempt:
            attempt = QuizAttempt.objects.create(
                quiz=quiz,
                student=request.user,
                attempt_number=submitted_attempts + 1,
                is_submitted=False,
            )

        answers_data = request.data.get("answers", [])
        serializer = QuizAnswerInputSerializer(data=answers_data, many=True)
        serializer.is_valid(raise_exception=True)
        answers_in = serializer.validated_data

        remaining = None
        if quiz.time_limit_minutes:
            elapsed = (timezone.now() - attempt.started_at).total_seconds()
            remaining = int((quiz.time_limit_minutes * 60) - elapsed)
        if remaining is not None and remaining <= 0:
            finalized = QuizTakeView()._auto_finalize_attempt(attempt)
            return Response(
                {
                    "detail": "Quiz time has ended and your attempt was auto-submitted.",
                    "attempt_id": str(finalized.id),
                    "score": float(finalized.score) if finalized.score is not None else None,
                    "max_score": float(finalized.max_score) if finalized.max_score is not None else None,
                    "pending_manual_grading": finalized.pending_manual_grading,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        questions = {str(q.id): q for q in QuizQuestion.objects.filter(quiz=quiz)}
        for item in answers_in:
            qid = str(item["question_id"])
            question = questions.get(qid)
            if not question:
                continue
            selected_choice = None
            if item.get("selected_choice_id"):
                selected_choice = QuizChoice.objects.filter(id=item["selected_choice_id"], question=question).first()
            QuizAnswer.objects.update_or_create(
                attempt=attempt,
                question=question,
                defaults={
                    "selected_choice": selected_choice,
                    "text_answer": item.get("text_answer"),
                },
            )

        finalized = QuizTakeView()._auto_finalize_attempt(attempt)
        _sync_student_activity_items(request.user)

        return Response(
            {
                "attempt_id": str(finalized.id),
                "score": float(finalized.score) if finalized.score is not None else None,
                "max_score": float(finalized.max_score) if finalized.max_score is not None else None,
                "pending_manual_grading": finalized.pending_manual_grading,
            }
        )


class QuizSaveProgressView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can save quiz progress."}, status=status.HTTP_403_FORBIDDEN)

        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)
        enrolled = Enrollment.objects.filter(course_section=quiz.course_section, student=request.user, is_active=True).exists()
        if not enrolled:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        attempt_id = request.data.get("attempt_id")
        attempt = QuizAttempt.objects.filter(id=attempt_id, quiz=quiz, student=request.user, is_submitted=False).first() if attempt_id else None
        if not attempt:
            return Response({"detail": "Open attempt not found."}, status=status.HTTP_404_NOT_FOUND)

        if quiz.time_limit_minutes:
            elapsed = (timezone.now() - attempt.started_at).total_seconds()
            if int((quiz.time_limit_minutes * 60) - elapsed) <= 0:
                finalized = QuizTakeView()._auto_finalize_attempt(attempt)
                return Response(
                    {
                        "detail": "Quiz time has ended and your attempt was auto-submitted.",
                        "attempt_id": str(finalized.id),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        answers_data = request.data.get("answers", [])
        serializer = QuizAnswerInputSerializer(data=answers_data, many=True)
        serializer.is_valid(raise_exception=True)
        answers_in = serializer.validated_data
        questions = {str(q.id): q for q in QuizQuestion.objects.filter(quiz=quiz)}

        for item in answers_in:
            qid = str(item["question_id"])
            question = questions.get(qid)
            if not question:
                continue
            selected_choice = None
            if item.get("selected_choice_id"):
                selected_choice = QuizChoice.objects.filter(id=item["selected_choice_id"], question=question).first()
            QuizAnswer.objects.update_or_create(
                attempt=attempt,
                question=question,
                defaults={
                    "selected_choice": selected_choice,
                    "text_answer": item.get("text_answer"),
                },
            )

        return Response({"ok": True, "attempt_id": str(attempt.id)})


class QuizMyLatestAttemptView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if request.user.role != User.Role.STUDENT:
            return Response({"detail": "Only students can view this endpoint."}, status=status.HTTP_403_FORBIDDEN)

        quiz = Quiz.objects.filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)

        attempts_used = QuizAttempt.objects.filter(quiz_id=pk, student=request.user, is_submitted=True).count()
        attempt = QuizAttempt.objects.filter(quiz_id=pk, student=request.user, is_submitted=True).order_by("-attempt_number").first()

        graded_attempts = QuizAttempt.objects.filter(quiz_id=pk, is_submitted=True, pending_manual_grading=False, score__isnull=False)
        total = graded_attempts.count()
        avg_score = graded_attempts.aggregate(avg=Avg("score"))["avg"]
        low_score = None
        high_score = None
        if total > 0:
            scores = [float(s.score) for s in graded_attempts if s.score is not None]
            if scores:
                low_score = min(scores)
                high_score = max(scores)
        rank = None
        percentile = None
        if attempt and total > 0 and attempt.score is not None:
            better_or_equal = graded_attempts.filter(score__gte=attempt.score).count()
            rank = better_or_equal
            below_or_equal = graded_attempts.filter(score__lte=attempt.score).count()
            percentile = (below_or_equal / total) * 100

        all_my_attempts = QuizAttempt.objects.filter(
            quiz_id=pk,
            student=request.user,
            is_submitted=True,
        ).order_by("-attempt_number")
        attempts_payload = []
        for a in all_my_attempts:
            duration_seconds = None
            if a.started_at and a.submitted_at:
                duration_seconds = int((a.submitted_at - a.started_at).total_seconds())
            attempts_payload.append(
                {
                    "attempt_number": a.attempt_number,
                    "score": float(a.score) if a.score is not None else None,
                    "max_score": float(a.max_score) if a.max_score is not None else None,
                    "pending_manual_grading": a.pending_manual_grading,
                    "submitted_at": a.submitted_at,
                    "duration_seconds": duration_seconds,
                }
            )

        return Response(
            {
                "attempt_id": str(attempt.id) if attempt else None,
                "score": float(attempt.score) if attempt and attempt.score is not None else None,
                "max_score": float(attempt.max_score) if attempt and attempt.max_score is not None else None,
                "pending_manual_grading": attempt.pending_manual_grading if attempt else False,
                "attempt_number": attempt.attempt_number if attempt else 0,
                "attempts_used": attempts_used,
                "attempt_limit": quiz.attempt_limit,
                "attempts_remaining": max(quiz.attempt_limit - attempts_used, 0),
                "class_stats": {
                    "graded_count": total,
                    "average_score": float(avg_score) if avg_score is not None else None,
                    "lowest_score": low_score,
                    "highest_score": high_score,
                    "rank": rank,
                    "percentile": round(percentile, 2) if percentile is not None else None,
                },
                "attempts": attempts_payload,
            }
        )


class QuizGradingListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)
        if request.user.role == User.Role.TEACHER and quiz.course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        attempts = QuizAttempt.objects.filter(quiz=quiz, is_submitted=True).select_related("student").order_by("-submitted_at")
        payload = []
        for a in attempts:
            row = {
                "quiz_id": str(quiz.id),
                "attempt_id": str(a.id),
                "student_id": str(a.student_id),
                "student_name": a.student.full_name,
                "score": float(a.score) if a.score is not None else None,
                "max_score": float(a.max_score) if a.max_score is not None else None,
                "pending_manual_grading": a.pending_manual_grading,
                "submitted_at": a.submitted_at,
                "answers": [],
            }
            answers = QuizAnswer.objects.filter(attempt=a).select_related("question", "selected_choice")
            for ans in answers:
                row["answers"].append(
                    {
                        "answer_id": str(ans.id),
                        "question_id": str(ans.question_id),
                        "question_text": ans.question.question_text,
                        "question_type": ans.question.question_type,
                        "points": float(ans.question.points),
                        "selected_choice_id": str(ans.selected_choice_id) if ans.selected_choice_id else None,
                        "selected_choice_text": ans.selected_choice.choice_text if ans.selected_choice else None,
                        "text_answer": ans.text_answer,
                        "is_correct": ans.is_correct,
                        "points_awarded": float(ans.points_awarded) if ans.points_awarded is not None else None,
                        "needs_manual_grading": ans.needs_manual_grading,
                    }
                )
            payload.append(row)
        return Response(payload)


class QuizAnswerGradeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        answer = QuizAnswer.objects.select_related("attempt__quiz__course_section", "question").filter(id=pk).first()
        if not answer:
            return Response({"detail": "Answer not found."}, status=status.HTTP_404_NOT_FOUND)

        course_section = answer.attempt.quiz.course_section
        if request.user.role == User.Role.TEACHER and course_section.teacher_id != request.user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        if request.user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        # Check if this was previously pending manual grading
        was_pending = answer.needs_manual_grading

        serializer = QuizAnswerGradeSerializer(answer, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        graded = serializer.save(
            needs_manual_grading=False,
            graded_at=timezone.now(),
            graded_by=request.user,
        )

        attempt = graded.attempt
        points_total = 0.0
        pending = False
        for ans in QuizAnswer.objects.filter(attempt=attempt):
            if ans.needs_manual_grading:
                pending = True
            if ans.points_awarded is not None:
                points_total += float(ans.points_awarded)
        attempt.score = points_total
        attempt.pending_manual_grading = pending
        attempt.save(update_fields=["score", "pending_manual_grading"])
        enrollment = Enrollment.objects.filter(
            course_section=attempt.quiz.course_section,
            student=attempt.student,
            is_active=True,
        ).first()
        if enrollment:
            _recompute_enrollment_grade(enrollment)

        # Send notification when grading is complete (no more pending)
        if was_pending and not pending:
            self._send_quiz_grade_notification(attempt)

        return Response(
            {
                "attempt_id": str(attempt.id),
                "score": float(attempt.score) if attempt.score is not None else None,
                "pending_manual_grading": attempt.pending_manual_grading,
            }
        )

    def _send_quiz_grade_notification(self, attempt: QuizAttempt):
        """Send push notification when quiz grading is complete."""
        from .push_notifications import send_push_notification_to_users

        try:
            quiz = attempt.quiz
            student = attempt.student

            # Create in-app notification
            Notification.objects.create(
                recipient=student,
                type=Notification.NotificationType.GRADE_RELEASED,
                title=f"Quiz Graded: {quiz.title}",
                body=f"Your quiz '{quiz.title}' has been graded. Score: {attempt.score}/{attempt.max_score}",
                course_section=quiz.course_section,
                quiz=quiz,
            )

            # Send push notification
            data = {
                "type": "grade_released",
                "quiz_id": str(quiz.id),
                "course_section_id": str(quiz.course_section_id),
            }

            send_push_notification_to_users(
                user_ids=[str(student.id)],
                title=f"Quiz Graded: {quiz.title}",
                body=f"Your quiz has been graded. Score: {attempt.score}/{attempt.max_score}",
                data=data,
            )
        except Exception as e:
            logger.warning(f"Failed to send quiz grade notification: {e}")


class QuizQuestionsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_teacher_access(self, request, quiz: Quiz):
        if request.user.role == User.Role.ADMIN:
            return None
        if request.user.role == User.Role.TEACHER and quiz.course_section.teacher_id == request.user.id:
            return None
        return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

    def get(self, request, pk):
        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)
        denied = self._ensure_teacher_access(request, quiz)
        if denied:
            return denied
        qs = QuizQuestion.objects.filter(quiz=quiz).prefetch_related("choices").order_by("sort_order")
        return Response(QuizQuestionWriteSerializer(qs, many=True).data)

    def post(self, request, pk):
        quiz = Quiz.objects.select_related("course_section").filter(id=pk).first()
        if not quiz:
            return Response({"detail": "Quiz not found."}, status=status.HTTP_404_NOT_FOUND)
        denied = self._ensure_teacher_access(request, quiz)
        if denied:
            return denied
        payload = dict(request.data)
        payload["quiz_id"] = str(quiz.id)
        serializer = QuizQuestionWriteSerializer(data=payload)
        serializer.is_valid(raise_exception=True)
        question = serializer.save()
        _recompute_course_section_grades(quiz.course_section)
        return Response(QuizQuestionWriteSerializer(question).data, status=status.HTTP_201_CREATED)


class QuizQuestionDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _resolve_question(self, pk):
        return QuizQuestion.objects.select_related("quiz__course_section").filter(id=pk).first()

    def _ensure_teacher_access(self, request, question: QuizQuestion):
        if request.user.role == User.Role.ADMIN:
            return None
        if request.user.role == User.Role.TEACHER and question.quiz.course_section.teacher_id == request.user.id:
            return None
        return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

    def patch(self, request, pk):
        question = self._resolve_question(pk)
        if not question:
            return Response({"detail": "Question not found."}, status=status.HTTP_404_NOT_FOUND)
        denied = self._ensure_teacher_access(request, question)
        if denied:
            return denied
        payload = dict(request.data)
        payload["quiz_id"] = str(question.quiz_id)
        serializer = QuizQuestionWriteSerializer(question, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        _recompute_course_section_grades(question.quiz.course_section)
        return Response(QuizQuestionWriteSerializer(question).data)

    def delete(self, request, pk):
        question = self._resolve_question(pk)
        if not question:
            return Response({"detail": "Question not found."}, status=status.HTTP_404_NOT_FOUND)
        denied = self._ensure_teacher_access(request, question)
        if denied:
            return denied
        course_section = question.quiz.course_section
        question.delete()
        _recompute_course_section_grades(course_section)
        return Response(status=status.HTTP_204_NO_CONTENT)


class QuizQuickCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        if user.role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        course_section_id = request.data.get("course_section_id")
        title = request.data.get("title")
        questions = request.data.get("questions") or []
        weekly_module_id = request.data.get("weekly_module_id")
        if not course_section_id or not title:
            return Response({"detail": "course_section_id and title are required."}, status=status.HTTP_400_BAD_REQUEST)

        course_section = CourseSection.objects.filter(id=course_section_id).first()
        if not course_section:
            return Response({"detail": "Course section not found."}, status=status.HTTP_404_NOT_FOUND)
        if user.role == User.Role.TEACHER and course_section.teacher_id != user.id:
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        weekly_module = None
        if weekly_module_id:
            weekly_module = WeeklyModule.objects.filter(id=weekly_module_id, course_section=course_section).first()
            if not weekly_module:
                return Response({"detail": "Selected week/topic is invalid for this course section."}, status=status.HTTP_400_BAD_REQUEST)

        def to_int(value, default):
            if value in [None, ""]:
                return default
            try:
                return int(value)
            except (TypeError, ValueError):
                return default

        quiz = Quiz.objects.create(
            course_section=course_section,
            weekly_module=weekly_module,
            title=title,
            instructions=request.data.get("instructions"),
            time_limit_minutes=to_int(request.data.get("time_limit_minutes"), None),
            attempt_limit=to_int(request.data.get("attempt_limit"), 1),
            open_at=request.data.get("open_at") or None,
            close_at=request.data.get("close_at") or None,
            is_published=request.data.get("is_published", True),
            shuffle_questions=request.data.get("shuffle_questions", False),
            shuffle_choices=request.data.get("shuffle_choices", False),
            show_results=request.data.get("show_results", True),
        )
        if quiz.is_published:
            _notify_students_for_course_section(
                course_section=quiz.course_section,
                notif_type=Notification.NotificationType.NEW_QUIZ,
                title=f"New Quiz: {quiz.title}",
                body=f"A new quiz was posted in {quiz.course_section.course.title}.",
                quiz=quiz,
            )

        created_questions = []
        for idx, q in enumerate(questions):
            q_payload = {
                "quiz_id": str(quiz.id),
                "question_text": q.get("question_text"),
                "question_type": q.get("question_type", QuizQuestion.QuestionType.MULTIPLE_CHOICE),
                "points": q.get("points", 1),
                "sort_order": q.get("sort_order", idx),
                "choices": q.get("choices", []),
            }
            s = QuizQuestionWriteSerializer(data=q_payload)
            s.is_valid(raise_exception=True)
            created_questions.append(s.save())

        _recompute_course_section_grades(quiz.course_section)
        _sync_course_section_students_activity_items(quiz.course_section)

        return Response(
            {
                "quiz": QuizSerializer(quiz).data,
                "questions": QuizQuestionWriteSerializer(created_questions, many=True).data,
            },
            status=status.HTTP_201_CREATED,
        )


class TeacherCourseSectionScopedModelViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    course_section_field = "course_section"

    def _get_course_section_id(self):
        payload = getattr(self.request, "data", {}) or {}
        return payload.get("course_section_id") or self.request.query_params.get("course_section_id")

    def _can_access_course_section(self, course_section: CourseSection) -> bool:
        user = self.request.user
        if user.role == User.Role.ADMIN:
            return True
        return user.role == User.Role.TEACHER and course_section.teacher_id == user.id

    def _scope_queryset(self, qs):
        user = self.request.user
        if user.role == User.Role.ADMIN:
            return qs
        if user.role == User.Role.TEACHER:
            lookup = {f"{self.course_section_field}__teacher": user}
            return qs.filter(**lookup)
        return qs.none()

    def get_queryset(self):
        return self._scope_queryset(self.queryset)

    def create(self, request, *args, **kwargs):
        course_section_id = self._get_course_section_id()
        if not course_section_id:
            return Response({"detail": "course_section_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        course_section = CourseSection.objects.filter(id=course_section_id).first()
        if not course_section:
            return Response({"detail": "Course section not found."}, status=status.HTTP_404_NOT_FOUND)
        if not self._can_access_course_section(course_section):
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def perform_update(self, serializer):
        instance = self.get_object()
        course_section = getattr(instance, self.course_section_field)
        if not self._can_access_course_section(course_section):
            raise permissions.PermissionDenied("Not allowed.")
        serializer.save()

    def perform_destroy(self, instance):
        course_section = getattr(instance, self.course_section_field)
        if not self._can_access_course_section(course_section):
            raise permissions.PermissionDenied("Not allowed.")
        instance.delete()


class WeeklyModuleViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = WeeklyModule.objects.all().order_by("week_number", "sort_order")
    serializer_class = WeeklyModuleSerializer

    def create(self, request, *args, **kwargs):
        return Response(
            {"detail": "Modules are auto-created from the course week count. Edit existing weeks instead."},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )


class AssignmentGroupViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = AssignmentGroup.objects.all().order_by("name")
    serializer_class = AssignmentGroupSerializer

    def perform_create(self, serializer):
        group = serializer.save()
        _recompute_course_section_grades(group.course_section)

    def perform_update(self, serializer):
        group = serializer.save()
        _recompute_course_section_grades(group.course_section)

    def perform_destroy(self, instance):
        course_section = instance.course_section
        super().perform_destroy(instance)
        _recompute_course_section_grades(course_section)


class ActivityViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = Activity.objects.all().order_by("deadline")
    serializer_class = ActivitySerializer

    def get_object(self):
        """Allow students to retrieve activities in their enrolled courses."""
        from django.shortcuts import get_object_or_404
        from rest_framework.exceptions import NotFound

        pk = self.kwargs.get('pk')
        user = self.request.user

        # Get the activity
        activity = Activity.objects.filter(id=pk).first()
        if not activity:
            raise NotFound("Activity not found.")

        # Check permissions based on role
        if user.role == User.Role.ADMIN:
            return activity

        if user.role == User.Role.TEACHER:
            if activity.course_section.teacher_id == user.id:
                return activity
            raise NotFound("Activity not found.")

        if user.role == User.Role.STUDENT:
            # Students can access if enrolled in the course section
            is_enrolled = Enrollment.objects.filter(
                student=user,
                course_section=activity.course_section,
                is_active=True
            ).exists()
            if is_enrolled and activity.is_published:
                return activity
            raise NotFound("Activity not found.")

        raise NotFound("Activity not found.")

    def retrieve(self, request, *args, **kwargs):
        """Allow students to view activities in their enrolled courses."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data

        # Add my_submission for students to enable dynamic UI updates after submission
        if request.user.role == User.Role.STUDENT:
            all_submissions = Submission.objects.filter(
                activity=instance,
                student=request.user
            ).order_by('-attempt_number')
            # Latest submission (first in the ordered list)
            latest_submission = all_submissions.first()
            data['my_submission'] = SubmissionSerializer(latest_submission).data if latest_submission else None
            # Include all submissions for attempt history dropdown
            data['my_submissions'] = SubmissionSerializer(all_submissions, many=True).data
            # Include attempt limits
            attempt_limit = instance.attempt_limit or 1
            attempts_used = all_submissions.count()
            data['attempt_limit'] = attempt_limit
            data['attempts_used'] = attempts_used
            data['attempts_remaining'] = max(attempt_limit - attempts_used, 0)

        return Response(data)

    def _validate_weekly_module(self, serializer):
        weekly_module = serializer.validated_data.get("weekly_module")
        course_section = serializer.validated_data.get("course_section") or getattr(serializer.instance, "course_section", None)
        if weekly_module and course_section and weekly_module.course_section_id != course_section.id:
            raise permissions.PermissionDenied("Selected week/topic does not belong to this course section.")
        assignment_group = serializer.validated_data.get("assignment_group")
        if assignment_group and course_section and assignment_group.course_section_id != course_section.id:
            raise permissions.PermissionDenied("Selected assignment group does not belong to this course section.")

    def perform_create(self, serializer):
        self._validate_weekly_module(serializer)
        activity = serializer.save(created_by=self.request.user)
        _recompute_course_section_grades(activity.course_section)
        _sync_course_section_students_activity_items(activity.course_section)
        if activity.is_published:
            _notify_students_for_course_section(
                course_section=activity.course_section,
                notif_type=Notification.NotificationType.NEW_ACTIVITY,
                title=f"New Assignment: {activity.title}",
                body=f"A new assignment was posted in {activity.course_section.course.title}.",
                activity=activity,
            )

    def perform_update(self, serializer):
        self._validate_weekly_module(serializer)
        before = serializer.instance
        old_published = bool(before.is_published)
        old_title = before.title
        old_deadline = before.deadline
        old_points = before.points
        old_instructions = before.instructions
        old_description = before.description

        updated = serializer.save()
        _recompute_course_section_grades(updated.course_section)
        _sync_course_section_students_activity_items(updated.course_section)

        became_published = (not old_published) and bool(updated.is_published)
        changed_while_published = bool(updated.is_published) and (
            updated.title != old_title
            or updated.deadline != old_deadline
            or updated.points != old_points
            or (updated.instructions or "") != (old_instructions or "")
            or (updated.description or "") != (old_description or "")
        )

        if became_published:
            _notify_students_for_course_section(
                course_section=updated.course_section,
                notif_type=Notification.NotificationType.NEW_ACTIVITY,
                title=f"New Assignment: {updated.title}",
                body=f"A new assignment was posted in {updated.course_section.course.title}.",
                activity=updated,
            )
        elif changed_while_published:
            _notify_students_for_course_section(
                course_section=updated.course_section,
                notif_type=Notification.NotificationType.NEW_ACTIVITY,
                title=f"Updated Assignment: {updated.title}",
                body=f"An assignment was updated in {updated.course_section.course.title}.",
                activity=updated,
            )

    def perform_destroy(self, instance):
        course_section = instance.course_section
        super().perform_destroy(instance)
        _recompute_course_section_grades(course_section)
        _sync_course_section_students_activity_items(course_section)


class CourseFileViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = CourseFile.objects.all().order_by("-created_at")
    serializer_class = CourseFileSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def _validate_weekly_module(self, serializer):
        weekly_module = serializer.validated_data.get("weekly_module")
        course_section = serializer.validated_data.get("course_section") or getattr(serializer.instance, "course_section", None)
        if weekly_module and course_section and weekly_module.course_section_id != course_section.id:
            raise permissions.PermissionDenied("Selected week/topic does not belong to this course section.")

    def perform_create(self, serializer):
        self._validate_weekly_module(serializer)
        serializer.save(uploader=self.request.user)

    def perform_update(self, serializer):
        self._validate_weekly_module(serializer)
        super().perform_update(serializer)

    def _inject_uploaded_file_data(self, request, payload):
        file_obj = request.FILES.get("file")
        if not file_obj:
            return payload

        path = default_storage.save(
            f"course_files/{request.user.id}/{timezone.now().timestamp()}_{file_obj.name}",
            file_obj,
        )
        file_url = request.build_absolute_uri(default_storage.url(path))
        payload["file_url"] = file_url
        payload["file_name"] = payload.get("file_name") or file_obj.name
        payload["file_size_bytes"] = file_obj.size
        ext = file_obj.name.split(".")[-1].lower() if "." in file_obj.name else None
        payload["file_type"] = payload.get("file_type") or ext
        preview_url = _convert_office_upload_to_pdf_preview(
            request=request,
            source_storage_path=path,
            original_name=file_obj.name,
        )
        if preview_url:
            payload["preview_file_url"] = preview_url
        return payload

    def create(self, request, *args, **kwargs):
        course_section_id = self._get_course_section_id()
        if not course_section_id:
            return Response({"detail": "course_section_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        course_section = CourseSection.objects.filter(id=course_section_id).first()
        if not course_section:
            return Response({"detail": "Course section not found."}, status=status.HTTP_404_NOT_FOUND)
        if not self._can_access_course_section(course_section):
            return Response({"detail": "Not allowed."}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data.copy()
        payload = self._inject_uploaded_file_data(request, payload)
        if not payload.get("file_url"):
            return Response({"detail": "Either file_url or an uploaded file is required."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        payload = request.data.copy()
        payload = self._inject_uploaded_file_data(request, payload)
        if not payload.get("file_url"):
            payload["file_url"] = instance.file_url

        serializer = self.get_serializer(instance, data=payload, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)


class AnnouncementViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = Announcement.objects.all().order_by("-created_at")
    serializer_class = AnnouncementSerializer

    def perform_create(self, serializer):
        announcement = serializer.save(created_by=self.request.user)

        # Send notifications to students if published
        if announcement.is_published:
            self._send_announcement_notifications(announcement)

    def perform_update(self, serializer):
        announcement = serializer.save()
        # Send notifications if being published for the first time
        if announcement.is_published:
            self._send_announcement_notifications(announcement)

    def _send_announcement_notifications(self, announcement: Announcement):
        """Send in-app notifications and push notifications for announcement."""
        from .push_notifications import send_push_notification_to_users

        # Determine recipients based on school_wide flag and audience
        if announcement.school_wide:
            # School-wide announcement - notify all teachers and students
            recipient_ids = list(
                User.objects.filter(
                    status=User.Status.ACTIVE,
                    role__in=[User.Role.TEACHER, User.Role.STUDENT],
                ).values_list('id', flat=True)
            )
            course_section_id = None
        else:
            # Course section announcement
            if not announcement.course_section:
                return
            course_section_id = str(announcement.course_section.id)

            # Filter recipients based on audience
            if announcement.audience == Announcement.Audience.TEACHERS_ONLY:
                recipient_ids = [announcement.course_section.teacher_id] if announcement.course_section.teacher else []
            else:
                # ALL - notify all enrolled students
                recipient_ids = list(
                    Enrollment.objects.filter(
                        course_section=announcement.course_section,
                        is_active=True,
                        student__status=User.Status.ACTIVE,
                    ).values_list('student_id', flat=True)
                )

        if not recipient_ids:
            return

        # Create in-app notifications
        notif_type = (
            Notification.NotificationType.SCHOOL_ANNOUNCEMENT
            if announcement.school_wide
            else Notification.NotificationType.COURSE_ANNOUNCEMENT
        )

        notifications = [
            Notification(
                recipient_id=recipient_id,
                type=notif_type,
                title=announcement.title,
                body=announcement.body[:200] if announcement.body else "",
                course_section=announcement.course_section,
                announcement=announcement,
            )
            for recipient_id in recipient_ids
        ]
        Notification.objects.bulk_create(notifications)

        # Send push notifications
        data = {
            "type": notif_type,
            "announcement_id": str(announcement.id),
        }
        if course_section_id:
            data["course_section_id"] = course_section_id

        try:
            send_push_notification_to_users(
                user_ids=[str(uid) for uid in recipient_ids],
                title=announcement.title,
                body=announcement.body[:100] if announcement.body else "",
                data=data,
            )
        except Exception as e:
            logger.warning(f"Failed to send push notifications for announcement: {e}")


class QuizViewSet(TeacherCourseSectionScopedModelViewSet):
    queryset = Quiz.objects.all().order_by("-created_at")
    serializer_class = QuizSerializer

    def get_object(self):
        """Allow students to retrieve quizzes in their enrolled courses."""
        from django.shortcuts import get_object_or_404
        from rest_framework.exceptions import NotFound

        pk = self.kwargs.get('pk')
        user = self.request.user

        # Get the quiz
        quiz = Quiz.objects.filter(id=pk).first()
        if not quiz:
            raise NotFound("Quiz not found.")

        # Check permissions based on role
        if user.role == User.Role.ADMIN:
            return quiz

        if user.role == User.Role.TEACHER:
            if quiz.course_section.teacher_id == user.id:
                return quiz
            raise NotFound("Quiz not found.")

        if user.role == User.Role.STUDENT:
            # Students can access if enrolled in the course section
            is_enrolled = Enrollment.objects.filter(
                student=user,
                course_section=quiz.course_section,
                is_active=True
            ).exists()
            if is_enrolled and quiz.is_published:
                return quiz
            raise NotFound("Quiz not found.")

        raise NotFound("Quiz not found.")

    def retrieve(self, request, *args, **kwargs):
        """Allow students to view quizzes in their enrolled courses."""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def _validate_weekly_module(self, serializer):
        weekly_module = serializer.validated_data.get("weekly_module")
        course_section = serializer.validated_data.get("course_section") or getattr(serializer.instance, "course_section", None)
        if weekly_module and course_section and weekly_module.course_section_id != course_section.id:
            raise permissions.PermissionDenied("Selected week/topic does not belong to this course section.")

    def perform_create(self, serializer):
        self._validate_weekly_module(serializer)
        quiz = serializer.save()
        _recompute_course_section_grades(quiz.course_section)
        _sync_course_section_students_activity_items(quiz.course_section)
        if quiz.is_published:
            _notify_students_for_course_section(
                course_section=quiz.course_section,
                notif_type=Notification.NotificationType.NEW_QUIZ,
                title=f"New Quiz: {quiz.title}",
                body=f"A new quiz was posted in {quiz.course_section.course.title}.",
                quiz=quiz,
            )

    def perform_update(self, serializer):
        self._validate_weekly_module(serializer)
        before = serializer.instance
        old_published = bool(before.is_published)
        old_title = before.title
        old_instructions = before.instructions
        old_attempt_limit = before.attempt_limit
        old_time_limit = before.time_limit_minutes
        old_open_at = before.open_at
        old_close_at = before.close_at

        updated = serializer.save()
        _recompute_course_section_grades(updated.course_section)
        _sync_course_section_students_activity_items(updated.course_section)

        became_published = (not old_published) and bool(updated.is_published)
        changed_while_published = bool(updated.is_published) and (
            updated.title != old_title
            or (updated.instructions or "") != (old_instructions or "")
            or updated.attempt_limit != old_attempt_limit
            or updated.time_limit_minutes != old_time_limit
            or updated.open_at != old_open_at
            or updated.close_at != old_close_at
        )

        if became_published:
            _notify_students_for_course_section(
                course_section=updated.course_section,
                notif_type=Notification.NotificationType.NEW_QUIZ,
                title=f"New Quiz: {updated.title}",
                body=f"A new quiz was posted in {updated.course_section.course.title}.",
                quiz=updated,
            )
        elif changed_while_published:
            _notify_students_for_course_section(
                course_section=updated.course_section,
                notif_type=Notification.NotificationType.NEW_QUIZ,
                title=f"Updated Quiz: {updated.title}",
                body=f"A quiz was updated in {updated.course_section.course.title}.",
                quiz=updated,
            )

    def perform_destroy(self, instance):
        course_section = instance.course_section
        super().perform_destroy(instance)
        _recompute_course_section_grades(course_section)
        _sync_course_section_students_activity_items(course_section)


class PushTokenViewSet(viewsets.ModelViewSet):
    """Viewset for managing push notification tokens."""
    serializer_class = PushTokenSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return PushToken.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        # If token already exists for another user, deactivate it
        token = serializer.validated_data.get('token')
        if token:
            PushToken.objects.filter(token=token).exclude(user=self.request.user).update(is_active=False)
        serializer.save(user=self.request.user)


class ActivityReminderViewSet(viewsets.ModelViewSet):
    """Viewset for managing activity/quiz reminders."""
    serializer_class = ActivityReminderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = ActivityReminder.objects.filter(user=self.request.user).select_related('activity', 'quiz')

        # Filter by activity_id or quiz_id if provided
        activity_id = self.request.query_params.get('activity_id')
        quiz_id = self.request.query_params.get('quiz_id')
        reminder_type = self.request.query_params.get('reminder_type')

        if activity_id:
            qs = qs.filter(activity_id=activity_id)
        if quiz_id:
            qs = qs.filter(quiz_id=quiz_id)
        if reminder_type:
            qs = qs.filter(reminder_type=reminder_type)

        return qs.order_by('reminder_datetime')

    def perform_create(self, serializer):
        # Validate that the user has access to the activity or quiz
        reminder_type = serializer.validated_data.get('reminder_type')
        activity = serializer.validated_data.get('activity')
        quiz = serializer.validated_data.get('quiz')

        if reminder_type == 'activity' and activity:
            # Check if user is enrolled in the course section
            enrolled = Enrollment.objects.filter(
                course_section=activity.course_section,
                student=self.request.user,
                is_active=True,
            ).exists()
            if not enrolled and self.request.user.role != User.Role.ADMIN:
                raise permissions.PermissionDenied("You are not enrolled in this activity's course section.")

        if reminder_type == 'quiz' and quiz:
            # Check if user is enrolled in the course section
            enrolled = Enrollment.objects.filter(
                course_section=quiz.course_section,
                student=self.request.user,
                is_active=True,
            ).exists()
            if not enrolled and self.request.user.role != User.Role.ADMIN:
                raise permissions.PermissionDenied("You are not enrolled in this quiz's course section.")

        serializer.save(user=self.request.user)


class PushTokenViewSet(viewsets.ModelViewSet):
    """Viewset for managing push notification tokens."""
    serializer_class = PushTokenSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return PushToken.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        # If token already exists for another user, deactivate it
        token = serializer.validated_data.get('token')
        if token:
            PushToken.objects.filter(token=token).exclude(user=self.request.user).update(is_active=False)
        serializer.save(user=self.request.user)


class ActivityReminderViewSet(viewsets.ModelViewSet):
    """Viewset for managing activity/quiz reminders."""
    serializer_class = ActivityReminderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = ActivityReminder.objects.filter(user=self.request.user).select_related('activity', 'quiz')

        # Filter by activity_id or quiz_id if provided
        activity_id = self.request.query_params.get('activity_id')
        quiz_id = self.request.query_params.get('quiz_id')
        reminder_type = self.request.query_params.get('reminder_type')

        if activity_id:
            qs = qs.filter(activity_id=activity_id)
        if quiz_id:
            qs = qs.filter(quiz_id=quiz_id)
        if reminder_type:
            qs = qs.filter(reminder_type=reminder_type)

        return qs.order_by('reminder_datetime')

    def perform_create(self, serializer):
        # Validate that the user has access to the activity or quiz
        reminder_type = serializer.validated_data.get('reminder_type')
        activity = serializer.validated_data.get('activity')
        quiz = serializer.validated_data.get('quiz')

        if reminder_type == 'activity' and activity:
            # Check if user is enrolled in the course section
            enrolled = Enrollment.objects.filter(
                course_section=activity.course_section,
                student=self.request.user,
                is_active=True,
            ).exists()
            if not enrolled and self.request.user.role != User.Role.ADMIN:
                raise permissions.PermissionDenied("You are not enrolled in this activity's course section.")

        if reminder_type == 'quiz' and quiz:
            # Check if user is enrolled in the course section
            enrolled = Enrollment.objects.filter(
                course_section=quiz.course_section,
                student=self.request.user,
                is_active=True,
            ).exists()
            if not enrolled and self.request.user.role != User.Role.ADMIN:
                raise permissions.PermissionDenied("You are not enrolled in this quiz's course section.")

        serializer.save(user=self.request.user)


class ActivityCommentViewSet(viewsets.ModelViewSet):
    """ViewSet for activity comments - allows students and teachers to comment on activities."""

    queryset = ActivityComment.objects.all().select_related('author', 'activity').prefetch_related('replies')
    serializer_class = ActivityCommentSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        """Filter comments by activity_id, optionally filter by submission_id."""
        qs = super().get_queryset()

        activity_id = self.request.query_params.get('activity_id')
        submission_id = self.request.query_params.get('submission_id')

        if activity_id:
            qs = qs.filter(activity_id=activity_id)
        if submission_id:
            qs = qs.filter(submission_id=submission_id)

        # Only return top-level comments (parent=None), replies are nested
        qs = qs.filter(parent=None)

        return qs.order_by('-created_at')

    def create(self, request, *args, **kwargs):
        """Create a new comment with optional file attachments."""
        activity_id = request.data.get('activity_id')
        if not activity_id:
            return Response({"detail": "activity_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        activity = Activity.objects.filter(id=activity_id).first()
        if not activity:
            return Response({"detail": "Activity not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check user access - student must be enrolled, teacher must teach the course
        if request.user.role == User.Role.STUDENT:
            enrolled = Enrollment.objects.filter(
                course_section=activity.course_section,
                student=request.user,
                is_active=True,
            ).exists()
            if not enrolled:
                return Response({"detail": "You are not enrolled in this activity's course section."}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role == User.Role.TEACHER:
            teaches = CourseSection.objects.filter(
                id=activity.course_section_id,
                teacher=request.user,
            ).exists()
            if not teaches:
                return Response({"detail": "You are not the teacher of this course section."}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role != User.Role.ADMIN:
            return Response({"detail": "Not authorized."}, status=status.HTTP_403_FORBIDDEN)

        # Handle file uploads
        file_urls = request.data.get('file_urls') or []
        if not isinstance(file_urls, list):
            file_urls = []

        # Process uploaded files
        files = request.FILES.getlist('files')
        for file_obj in files:
            path = default_storage.save(
                f"comments/{request.user.id}/{timezone.now().timestamp()}_{file_obj.name}",
                file_obj,
            )
            file_urls.append(request.build_absolute_uri(default_storage.url(path)))

        content = request.data.get('content')
        parent_id = request.data.get('parent_id')
        submission_id = request.data.get('submission_id')

        # Debug logging for submission_id
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[ActivityComment] Creating comment - activity_id: {activity_id}, submission_id: {submission_id}, author: {request.user.id}")

        parent_comment = None
        if parent_id:
            parent_comment = ActivityComment.objects.filter(id=parent_id).first()
            if not parent_comment:
                return Response({"detail": "Parent comment not found."}, status=status.HTTP_404_NOT_FOUND)
            # Parent must be for the same activity
            if str(parent_comment.activity_id) != str(activity_id):
                return Response({"detail": "Parent comment must be for the same activity."}, status=status.HTTP_400_BAD_REQUEST)

        comment = ActivityComment.objects.create(
            activity=activity,
            submission_id=submission_id,
            author=request.user,
            parent=parent_comment,
            content=content,
            file_urls=file_urls if file_urls else None,
        )

        logger.info(f"[ActivityComment] Comment created successfully - id: {comment.id}, submission_id: {comment.submission_id}, activity_id: {activity_id}")

        serializer = self.get_serializer(comment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        """Update a comment - only the author can update."""
        comment = self.get_object()
        if comment.author != request.user:
            return Response({"detail": "You can only edit your own comments."}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete a comment - only the author can delete."""
        comment = self.get_object()
        if comment.author != request.user:
            return Response({"detail": "You can only delete your own comments."}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


class ActivityCommentsByActivityView(APIView):
    """Get comments for a specific activity with per-user privacy.

    For students: Only returns comments they are involved in:
    - Comments authored by the student
    - Comments authored by teachers of the course section
    - Comments on their own submissions

    For teachers: Returns all comments for activities in their course sections.

    For admins: Returns all comments.
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        """Get comments for an activity based on user role and involvement."""
        activity = Activity.objects.filter(id=pk).first()
        if not activity:
            return Response({"detail": "Activity not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get the course section teacher for this activity
        course_section = activity.course_section
        teacher_id = course_section.teacher_id if course_section else None

        if request.user.role == User.Role.STUDENT:
            # Check enrollment
            enrolled = Enrollment.objects.filter(
                course_section=activity.course_section,
                student=request.user,
                is_active=True,
            ).exists()
            if not enrolled:
                return Response({"detail": "You are not enrolled in this activity's course section."}, status=status.HTTP_403_FORBIDDEN)

            # Get the student's submission for this activity (if any)
            student_submission = Submission.objects.filter(
                activity=activity,
                student=request.user,
            ).order_by('-submitted_at').first()
            student_submission_id = student_submission.id if student_submission else None

            # Build filter for comments visible to this student:
            # 1. Comments authored by this student
            # 2. Comments authored by the course section teacher
            # 3. Comments specifically on this student's submission (if they have one)
            from django.db.models import Q

            # Base filter: comments by this student OR by the teacher
            visible_filter = Q(author=request.user) | Q(author_id=teacher_id)

            # If student has a submission, also include comments on that submission
            if student_submission_id:
                visible_filter = visible_filter | Q(submission_id=student_submission_id)

            comments = ActivityComment.objects.filter(
                activity=activity,
                parent=None,
            ).filter(visible_filter).select_related('author').prefetch_related('replies__author').order_by('created_at')

            # Filter replies similarly - only include replies the student should see
            filtered_comments = []
            for comment in comments:
                filtered_comment = self._filter_comment_replies(comment, request.user.id, teacher_id, student_submission_id)
                if filtered_comment:
                    filtered_comments.append(filtered_comment)

            serializer = ActivityCommentSerializer(filtered_comments, many=True, context={'request': request})
            return Response(serializer.data)

        elif request.user.role == User.Role.TEACHER:
            # Check if teacher owns this course section
            teaches = CourseSection.objects.filter(
                id=activity.course_section_id,
                teacher=request.user,
            ).exists()
            if not teaches:
                return Response({"detail": "You are not the teacher of this course section."}, status=status.HTTP_403_FORBIDDEN)

            # Support filtering by submission_id (for submitted work) or student_id (for any student)
            submission_id = request.query_params.get('submission_id')
            student_id = request.query_params.get('student_id')

            # Debug logging
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"[ActivityComments] Teacher {request.user.id} fetching comments for activity {pk}, submission_id: {submission_id}, student_id: {student_id}")

            from django.db.models import Q

            if submission_id:
                # Filter by specific submission (for students who have submitted)
                # Verify the submission belongs to this activity
                submission = Submission.objects.filter(id=submission_id, activity=activity).first()
                if not submission:
                    logger.warning(f"[ActivityComments] Submission {submission_id} not found for activity {pk}")
                    return Response({"detail": "Submission not found for this activity."}, status=status.HTTP_404_NOT_FOUND)

                # Return comments for this specific submission
                # This includes: comments on this submission, and any back-and-forth
                comments = ActivityComment.objects.filter(
                    activity=activity,
                    submission_id=submission_id,
                    parent=None,
                ).select_related('author').prefetch_related('replies__author').order_by('created_at')
                logger.info(f"[ActivityComments] Found {comments.count()} comments for submission {submission_id}")

            elif student_id:
                # Filter by specific student (works even without submission)
                # Show conversation between this student and the teacher
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"[ActivityComments] Filtering by student_id: {student_id}, teacher_id: {request.user.id}")

                student_submissions = Submission.objects.filter(activity=activity, student_id=student_id)
                submission_ids = [s.id for s in student_submissions]
                logger.info(f"[ActivityComments] Found {len(submission_ids)} submissions for student")

                # Build filter: comments by this student OR by teacher OR on student's submissions
                student_filter = Q(author_id=student_id) | Q(author_id=request.user.id)
                if submission_ids:
                    student_filter = student_filter | Q(submission_id__in=submission_ids)

                # Log the filter conditions
                logger.info(f"[ActivityComments] Filter: author_id={student_id} OR author_id={request.user.id}" +
                           (f" OR submission_id in {submission_ids}" if submission_ids else ""))

                comments = ActivityComment.objects.filter(
                    activity=activity,
                    parent=None,
                ).filter(student_filter).select_related('author').prefetch_related('replies__author').order_by('created_at')

                logger.info(f"[ActivityComments] Found {comments.count()} top-level comments for student {student_id}")

                # Also filter replies to only show student-teacher conversation
                filtered_comments = []
                for comment in comments:
                    logger.info(f"[ActivityComments] Comment {comment.id}: author_id={comment.author_id}, submission_id={comment.submission_id}")
                    filtered_comment = self._filter_comment_replies_for_teacher(comment, student_id, request.user.id, submission_ids)
                    if filtered_comment:
                        filtered_comments.append(filtered_comment)

                logger.info(f"[ActivityComments] Returning {len(filtered_comments)} comments for student {student_id}")

                serializer = ActivityCommentSerializer(filtered_comments, many=True, context={'request': request})
                return Response(serializer.data)

            else:
                # Teachers see all comments for their course sections
                comments = ActivityComment.objects.filter(
                    activity=activity,
                    parent=None,
                ).select_related('author').prefetch_related('replies__author').order_by('created_at')

            serializer = ActivityCommentSerializer(comments, many=True, context={'request': request})
            return Response(serializer.data)

        elif request.user.role == User.Role.ADMIN:
            # Admins see all comments
            comments = ActivityComment.objects.filter(
                activity=activity,
                parent=None,
            ).select_related('author').prefetch_related('replies__author').order_by('created_at')

            serializer = ActivityCommentSerializer(comments, many=True, context={'request': request})
            return Response(serializer.data)

        return Response({"detail": "Not authorized."}, status=status.HTTP_403_FORBIDDEN)

    def _filter_comment_replies(self, comment, student_id, teacher_id, submission_id):
        """Filter replies to only include those the student should see."""
        visible_replies = []
        for reply in comment.replies.all():
            # Show reply if:
            # 1. Reply is by the student
            # 2. Reply is by the teacher
            # 3. Reply is on the student's submission (if they have one)
            is_by_student = str(reply.author_id) == str(student_id)
            is_by_teacher = str(reply.author_id) == str(teacher_id)
            is_on_student_submission = submission_id and str(reply.submission_id) == str(submission_id)

            if is_by_student or is_by_teacher or is_on_student_submission:
                visible_replies.append(reply)

        # Set filtered replies
        comment.replies._data = visible_replies
        return comment

    def _filter_comment_replies_for_teacher(self, comment, student_id, teacher_id, submission_ids):
        """Filter replies to only show student-teacher conversation."""
        visible_replies = []
        for reply in comment.replies.all():
            # Show reply if:
            # 1. Reply is by the student
            # 2. Reply is by the teacher
            # 3. Reply is on one of the student's submissions
            is_by_student = str(reply.author_id) == str(student_id)
            is_by_teacher = str(reply.author_id) == str(teacher_id)
            is_on_student_submission = submission_ids and str(reply.submission_id) in [str(sid) for sid in submission_ids]

            if is_by_student or is_by_teacher or is_on_student_submission:
                visible_replies.append(reply)

        # Set filtered replies
        comment.replies._data = visible_replies
        return comment

    def _filter_comment_replies_for_teacher(self, comment, student_id, teacher_id, submission_ids):
        """Filter replies to only show student-teacher conversation."""
        visible_replies = []
        for reply in comment.replies.all():
            # Show reply if:
            # 1. Reply is by the student
            # 2. Reply is by the teacher
            # 3. Reply is on one of the student's submissions
            is_by_student = str(reply.author_id) == str(student_id)
            is_by_teacher = str(reply.author_id) == str(teacher_id)
            is_on_student_submission = submission_ids and str(reply.submission_id) in [str(sid) for sid in submission_ids]

            if is_by_student or is_by_teacher or is_on_student_submission:
                visible_replies.append(reply)

        # Set filtered replies
        comment.replies._data = visible_replies
        return comment
